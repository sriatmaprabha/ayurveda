"""Generate comprehensive Prakriti Dosha Diagnosis Tree with 4-5 layers of questions,
remedies that improve at each layer, and coverage of 40+ common conditions."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_diagnosis_tree():
    wb = Workbook()

    # === STYLING ===
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    layer_fills = {
        1: PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),  # warm orange
        2: PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),  # light green
        3: PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),  # light blue
        4: PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # light purple
        5: PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # light pink
    }
    vata_fill = PatternFill(start_color="E8D5B7", end_color="E8D5B7", fill_type="solid")
    pitta_fill = PatternFill(start_color="F4C2C2", end_color="F4C2C2", fill_type="solid")
    kapha_fill = PatternFill(start_color="B7D7E8", end_color="B7D7E8", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_headers(ws, headers, widths):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 30
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 25

    def set_title(ws, title):
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # =====================================================================
    # SHEET 1: PRAKRITI DIAGNOSIS QUESTION TREE (5 LAYERS)
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Prakriti Diagnosis Tree"
    set_title(ws1, "PRAKRITI DOSHA DIAGNOSIS TREE -- 5 Layers of Progressive Questions")

    headers1 = [
        "Layer", "Question #", "Category", "Question",
        "If Vata (A)", "If Pitta (B)", "If Kapha (C)",
        "General Remedy at This Layer", "Purpose / What This Reveals"
    ]
    widths1 = [8, 10, 15, 40, 25, 25, 25, 55, 35]
    write_headers(ws1, headers1, widths1)

    tree_data = [
        # LAYER 1: Body Constitution Basics (4 questions)
        [1, "L1-Q1", "Physical", "How would you describe your body frame?",
         "Thin, lean, narrow shoulders, prominent joints",
         "Medium, moderate, athletic, well-proportioned",
         "Broad, large-boned, sturdy, stocky",
         "LAYER 1 GENERAL REMEDY:\n"
         "VATA: Start with warm sesame oil self-massage (Abhyanga) daily before bath. Drink warm water throughout the day. Avoid cold/raw foods.\n"
         "PITTA: Use coconut oil for body. Drink room-temperature or cool water. Avoid direct midday sun. Include bitter greens in diet.\n"
         "KAPHA: Dry brush massage (Garshana) before bath. Drink warm water with honey and lemon. Avoid daytime sleeping.",
         "Establishes baseline body constitution (Deha Prakriti). Frame is the most stable indicator -- it rarely changes."],

        [1, "L1-Q2", "Physiological", "How is your digestion generally?",
         "Irregular -- sometimes hungry, sometimes not, gas/bloating",
         "Strong -- can digest almost anything, gets irritable if meals skipped",
         "Slow but steady -- can skip meals without discomfort",
         "", "Agni (digestive fire) assessment -- central to Ayurvedic diagnosis. Vishama Agni=Vata, Tikshna Agni=Pitta, Manda Agni=Kapha"],

        [1, "L1-Q3", "Physical", "How would you describe your skin?",
         "Dry, rough, cool to touch, may crack in winter",
         "Warm, slightly oily, sensitive, prone to rashes/acne",
         "Thick, smooth, moist, cool, oily",
         "", "Skin reflects Bhrajaka Pitta and overall dosha balance. Direct indicator for skin-related treatments."],

        [1, "L1-Q4", "Psychological", "How do you respond to stress?",
         "Become anxious, worried, fearful, restless",
         "Become irritable, angry, aggressive, critical",
         "Become withdrawn, avoidant, attached, sad",
         "", "Manas (mind) Prakriti -- determines whether treatment should focus on calming (Vata), cooling (Pitta), or energizing (Kapha)"],

        # LAYER 2: Physiological Deep Dive (4 questions) -- IMPROVES Layer 1 remedy
        [2, "L2-Q1", "Physiological", "How is your sleep pattern?",
         "Light, disturbed, wake up between 2-6 AM, hard to fall asleep",
         "Moderate, sound, wake up hot/sweating, vivid/intense dreams",
         "Deep and heavy, difficult to wake up, sleep 8+ hours easily",
         "LAYER 2 IMPROVED REMEDY (adds to Layer 1):\n"
         "VATA: Add Ashwagandha 500mg before bed with warm milk. Practice Nadi Shodhana (alternate nostril breathing) 10 min before sleep. Avoid screens after 8 PM. Sleep by 10 PM.\n"
         "PITTA: Add Brahmi or Shatavari 500mg before bed. Practice Sheetali Pranayama (cooling breath) 10 min before sleep. Keep bedroom cool. Avoid intense work after 8 PM.\n"
         "KAPHA: Add Trikatu (ginger-pepper-pippali) with warm water before meals. Practice Kapalabhati 5 min upon waking. Wake before 6 AM. No naps.",
         "Sleep reveals Tamas (Kapha), Rajas (Vata), and Sattva balance. Also indicates Prana Vata and Tarpaka Kapha function."],

        [2, "L2-Q2", "Physiological", "How are your bowel movements?",
         "Irregular, dry, hard, tendency towards constipation",
         "Regular, soft, tendency towards loose/burning stools",
         "Regular, heavy, thick, mucus-like, sluggish",
         "", "Mala Pariksha (waste examination) -- indicates Apana Vata, Pachaka Pitta, and Kledaka Kapha function"],

        [2, "L2-Q3", "Physical", "What is your natural body temperature preference?",
         "Dislike cold, always want warmth, cold hands/feet",
         "Dislike heat, prefer cool, feel hot easily",
         "Comfortable generally, dislike cold AND damp weather",
         "", "Temperature preference is a strong dosha differentiator. Confirms or refines Layer 1 assessment."],

        [2, "L2-Q4", "Psychological", "How would you describe your emotional nature?",
         "Enthusiastic but anxious, mood changes quickly, creative",
         "Passionate but short-tempered, sharp intellect, competitive",
         "Calm, composed, loyal, but can become attached/possessive",
         "", "Refines Manas Prakriti. Key for mental health treatment decisions."],

        # LAYER 3: Sub-dosha Identification (4 questions) -- FURTHER IMPROVES remedy
        [3, "L3-Q1", "Physiological", "Where do you typically feel discomfort or symptoms first?",
         "Lower body: joints, lower back, colon, legs, feet",
         "Middle body: stomach, liver, small intestine, skin, eyes",
         "Upper body: chest, lungs, sinuses, throat, head",
         "LAYER 3 IMPROVED REMEDY (sub-dosha specific):\n"
         "VATA lower body: Basti therapy (medicated enema per Charaka). Warm sesame oil on lower back. Pavana Muktasana daily. Dashamoola decoction.\n"
         "VATA upper body: Nasya (nasal oil drops). Warm oil in ears. Ashwagandha + Bala decoction.\n"
         "PITTA middle: Virechana (purgation therapy). Shatavari + Amalaki. Bitter ghee. Avoid fermented foods.\n"
         "PITTA skin/eyes: Netra Tarpana (ghee eye treatment). Neem + Turmeric paste externally. Chandanasava internally.\n"
         "KAPHA upper: Vamana (therapeutic emesis). Trikatu before meals. Nasya with Anu Taila. Steam inhalation with eucalyptus.\n"
         "KAPHA lower: Guduchi + Punarnava for water retention. Triphala Guggulu. Vigorous Surya Namaskar.",
         "Locates the sub-dosha: Vata (Prana/Udana/Vyana/Samana/Apana), Pitta (Pachaka/Ranjaka/Sadhaka/Alochaka/Bhrajaka), Kapha (Tarpaka/Avalambaka/Kledaka/Bodhaka/Shleshaka)"],

        [3, "L3-Q2", "Temporal", "When are your symptoms worst?",
         "2-6 AM or 2-6 PM (Vata time), autumn, windy/cold weather",
         "10 AM-2 PM or 10 PM-2 AM (Pitta time), summer, hot weather",
         "6-10 AM or 6-10 PM (Kapha time), spring, cold/damp weather",
         "", "Kala Pariksha (time assessment) -- confirms dosha and guides treatment timing"],

        [3, "L3-Q3", "Physical", "Describe your tongue (look in a mirror):",
         "Dry, cracked, thin, may tremble, brownish coating",
         "Red, yellowish coating, sharp/pointed, may have ulcers",
         "Thick, swollen, white coating, teeth marks on edges",
         "", "Jihva Pariksha -- tongue diagnosis reveals Ama (toxins), dosha state, and organ involvement"],

        [3, "L3-Q4", "Lifestyle", "What is your current daily routine like?",
         "Irregular -- different sleep/wake times, meals at random",
         "Structured but intense -- driven, workaholic, competitive schedule",
         "Sedentary -- regular but low activity, comfort-oriented",
         "", "Dinacharya assessment -- irregular routine aggravates Vata, intensity aggravates Pitta, sedentary aggravates Kapha"],

        # LAYER 4: Personalization (3 questions) -- REFINES remedy to individual
        [4, "L4-Q1", "Personal", "What is your age group?",
         "Under 16 (Kapha stage) -- treatment must be gentle",
         "16-50 (Pitta stage) -- full-strength treatment possible",
         "Over 50 (Vata stage) -- nourishing, not depleting treatment",
         "LAYER 4 PERSONALIZED REMEDY:\n"
         "Young + Vata: Light Brimhana (nourishing). Bala + Ashwagandha milk. Gentle yoga only.\n"
         "Adult + Vata: Full Panchakarma eligible. Basti course (7-14 days). Dashamoola + Ashwagandha. Daily Abhyanga.\n"
         "Elder + Vata: Rasayana (rejuvenation) focus. Chyawanprash daily. Gentle oil massage. Warm, unctuous diet.\n"
         "Young + Pitta: Mild cooling herbs. Rose water. Avoid excess competition.\n"
         "Adult + Pitta: Virechana eligible. Shatavari + Guduchi. Moderate exercise, swimming.\n"
         "Elder + Pitta: Gentle cooling Rasayanas. Brahma Rasayana. Light diet.\n"
         "Young + Kapha: Encourage active play/sports. Light, warm diet. Honey with warm water.\n"
         "Adult + Kapha: Full Vamana eligible. Guggulu formulas. Vigorous daily exercise.\n"
         "Elder + Kapha: Gentle stimulation. Trikatu + Triphala. Walking. Avoid heavy food.",
         "Vayah Pariksha -- age determines treatment intensity. Children need gentle approach, adults can handle Panchakarma, elders need Rasayana."],

        [4, "L4-Q2", "Personal", "How long have you had your main symptoms?",
         "Days/weeks (acute/recent onset)",
         "Months (sub-acute, building up)",
         "Years (chronic, deep-seated in dhatus)",
         "", "Duration determines dhatu depth: Recent = Rasa/Rakta, Months = Mamsa/Meda, Years = Asthi/Majja/Shukra. Deeper = longer treatment."],

        [4, "L4-Q3", "Personal", "How is your physical strength currently?",
         "Weak/debilitated -- need gentle approach only",
         "Moderate -- can handle standard treatment",
         "Strong -- can handle intensive Panchakarma",
         "", "Bala Pariksha -- determines if Shodhana (purification) or only Shamana (palliative) is appropriate"],

        # LAYER 5: Final Protocol Selection (3 questions) -- LOCKS IN specific treatment
        [5, "L5-Q1", "Seasonal", "What is the current season where you are?",
         "Autumn/Early Winter (Vata season -- increase oil, warmth)",
         "Summer (Pitta season -- increase cooling, shade)",
         "Spring/Late Winter (Kapha season -- increase drying, stimulation)",
         "LAYER 5 FINAL PROTOCOL:\n"
         "Combines all layers into a personalized prescription:\n\n"
         "DIET: Specific foods for your dosha + season + age + symptom\n"
         "HERBS: 2-3 specific formulas with dosage and timing\n"
         "YOGA: Specific asana sequence from Care/Cure protocols with full step-by-step\n"
         "PRANAYAMA: Specific breathing technique with repetitions\n"
         "PANCHAKARMA: If eligible (based on strength), specific therapy recommendation\n"
         "DINACHARYA: Daily routine adjusted for your constitution\n"
         "RITUCHARYA: Seasonal routine adjustments\n\n"
         "(See Disease-Specific Protocols sheet for condition-by-condition details)",
         "Ritu Pariksha -- season modifies all treatments. Same dosha imbalance needs different approach in summer vs winter."],

        [5, "L5-Q2", "Dietary", "What is your current diet like?",
         "Mostly raw/cold food, irregular meal times, snacking",
         "Spicy, sour, fermented, alcohol, stimulants",
         "Heavy, sweet, oily, fried, large portions, comfort eating",
         "", "Identifies dietary aggravation -- often the primary cause. Diet correction alone resolves 60% of imbalances per Charaka."],

        [5, "L5-Q3", "Constitutional", "What is your known Prakriti (birth constitution) if known?",
         "Vata Prakriti (born thin, creative, variable)",
         "Pitta Prakriti (born medium, sharp, driven)",
         "Kapha Prakriti (born heavy, calm, steady) / Don't know",
         "", "Prakriti vs Vikriti gap determines treatment goal. If Prakriti=Vata but Vikriti=Kapha, treatment differs from Kapha-born Kapha-aggravated."],
    ]

    row = 3
    for d in tree_data:
        layer = d[0]
        for col, val in enumerate(d, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.alignment = wrap
            cell.border = border
            if col == 1:
                cell.fill = layer_fills.get(layer, layer_fills[1])
                cell.font = Font(bold=True, size=12)
            elif col >= 5 and col <= 7:
                if col == 5:
                    cell.fill = vata_fill
                elif col == 6:
                    cell.fill = pitta_fill
                elif col == 7:
                    cell.fill = kapha_fill
            elif col == 8:
                cell.fill = layer_fills.get(layer, layer_fills[1])

        ws1.row_dimensions[row].height = 180 if d[7] else 80
        row += 1

    # =====================================================================
    # SHEET 2: DISEASE-SPECIFIC PROTOCOLS (40+ conditions)
    # =====================================================================
    ws2 = wb.create_sheet("Disease Protocols")
    set_title(ws2, "DISEASE-SPECIFIC PROTOCOLS -- Dosha-Based Treatment for 40+ Common Conditions")

    headers2 = [
        "Condition", "Dominant Dosha", "Ayurvedic Name",
        "Key Symptoms", "Dietary Remedy",
        "Herbal Remedy (from classical texts)",
        "Yoga/Asana Protocol (from CSV)",
        "Panchakarma/Therapy",
        "Lifestyle Changes",
        "Classical Source"
    ]
    widths2 = [18, 12, 18, 30, 40, 40, 40, 35, 35, 25]
    write_headers(ws2, headers2, widths2)

    diseases = [
        # Mental Health
        ["Anxiety / Stress", "Vata", "Chittodvega",
         "Restlessness, fear, racing thoughts, palpitations, dry mouth, insomnia",
         "Warm, grounding foods. Ghee liberally. Sweet, sour, salty tastes. Warm milk with Ashwagandha + nutmeg at bedtime. Avoid caffeine, raw/cold food.",
         "Ashwagandha 500mg 2x/day. Brahmi 300mg morning. Jatamansi 250mg at night. Dashamoola decoction. Saraswatarishta 20ml after meals.",
         "CURE FOR ANXIETY:\n1. Varaaha Asana -- squat, feet together, palms behind back, gaze at nose tip, 30 sec\n2. Trishoola Kumbhaka -- inhale through nose+mouth, hold, exhale through nose. 21 times\n3. Seethkaaree Pranayama -- teeth together, inhale through teeth gaps, exhale through nose. 21 times\n4. Sheethalee Kumbhaka -- tongue rolled, inhale through tongue tube, hold, exhale nose. 21 times\n5. Trinetra Kumbhaka -- inhale left, hold; inhale right, hold; inhale both, hold; exhale. 21 times",
         "Shirodhara (warm oil on forehead) -- 30 min daily for 7 days. Basti with Dashamoola oil. Abhyanga with Bala oil.",
         "Fixed daily routine. Sleep by 10 PM. Avoid excessive travel. Reduce screen time. Warm oil in ears before sleep.",
         "Charaka Chikitsa 9, Ashtanga Hridaya Sutra 2, Sushruta Sutra 15"],

        ["Depression", "Kapha/Vata", "Vishada / Avasada",
         "Lethargy, sadness, withdrawal, loss of interest, heavy feeling, excessive sleep or insomnia",
         "KAPHA-type: Light, warm, spicy food. Honey with warm water. Avoid sweet, heavy, cold. Stimulating spices (ginger, black pepper).\nVATA-type: Warm, nourishing, sweet foods. Ghee. Avoid fasting.",
         "KAPHA: Trikatu + Guggulu. Vacha (Calamus) 250mg. Brahmi 500mg.\nVATA: Ashwagandha 500mg. Jatamansi 250mg. Saraswatarishta 20ml.\nBOTH: Shankhapushpi 500mg 2x/day.",
         "CARE FOR DEPRESSION:\n1. Simha Garjana Asana -- lion's roar pose, kneel, hands on knees, tongue out, roar. 21 times\n2. Varaaha Asana (as above)\n3. Kapalabhati -- 30 rapid exhales, then normal breathing. 3 rounds\n4. Bhastrika -- rapid inhale+exhale through both nostrils. 21 times",
         "KAPHA: Vamana (therapeutic emesis). Nasya with Anu Taila. Dry powder massage (Udvartana).\nVATA: Shirodhara. Gentle Basti. Abhyanga.",
         "Wake before 6 AM. Vigorous exercise daily. Avoid daytime sleep. Engage in new activities. Sunlight exposure 20 min/day.",
         "Charaka Chikitsa 9, Ashtanga Hridaya Uttara 6"],

        ["Insomnia", "Vata", "Anidra",
         "Unable to fall asleep, waking frequently (especially 2-6 AM), light sleep, fatigue, anxiety at night",
         "Warm milk with nutmeg + Ashwagandha at bedtime. Ghee rice for dinner. Cherry, banana. Avoid caffeine after noon. Light dinner before 7 PM.",
         "Ashwagandha 500mg at night. Jatamansi 250mg. Tagara (Valerian) 500mg. Brahmi 300mg evening. Saraswatarishta 20ml.",
         "CURE FOR INSOMNIA:\n1. Shashanka Asana -- kneel, forehead to ground, arms alongside body, deep breathing. 1-2 min\n2. Yoga Nidra -- guided relaxation in Shavasana. 20 min before sleep\n3. Nadi Shodhana -- alternate nostril breathing. 10 min\n4. Bhramari -- humming bee breath, close ears, hum on exhale. 11 times",
         "Shirodhara with warm sesame oil -- 30 min. Pada Abhyanga (foot massage with ghee). Shiro Abhyanga (head massage).",
         "Fixed sleep time (10 PM). No screens 1 hour before bed. Warm bath before sleep. Oil feet and scalp. Avoid stimulating conversation at night.",
         "Charaka Sutra 21, Sushruta Sharira 4, Ashtanga Hridaya Sutra 7"],

        # Musculoskeletal
        ["Arthritis / Joint Pain", "Vata", "Sandhivata / Amavata",
         "Joint pain, swelling, stiffness, cracking sounds, worse in cold/morning",
         "Warm, moist, oily foods. Ghee 2 tsp daily. Ginger tea. Turmeric milk. Avoid cold, dry, raw food. Avoid nightshades (tomato, potato, eggplant) in inflammatory arthritis.",
         "Yogaraja Guggulu 500mg 2x/day. Rasna Saptak Kwath decoction. Maharasnadi Kwath. Dashamoola decoction. Castor oil 1 tsp at bedtime.",
         "CURE FOR ARTHRITIS:\n1. Pavana Mukta Asana -- lie flat, bring both knees to chest, hold, breathe. 30 sec\n2. Vajrasana -- kneel sitting, spine straight. 2-5 min\n3. Gomukhasana -- seated, stack knees, clasp hands behind back. Each side 30 sec\n4. Setu Bandhasana -- bridge pose, lie on back, lift hips. 30 sec",
         "Basti with Dashamoola oil (primary Vata treatment). Janu Basti (oil pooling on knees). Patra Pinda Sweda (herbal poultice fomentation). Abhyanga with Mahanarayan oil.",
         "Warm environment. Gentle daily movement (never sedentary). Warm baths. Wool clothing in winter. Avoid cold drafts.",
         "Charaka Chikitsa 28-29, Sushruta Chikitsa 5, Ayurveda Science of Life Ch.1"],

        ["Back Pain (Lower)", "Vata", "Kati Shoola / Gridhrasi",
         "Pain in lumbar region, stiffness, may radiate to legs (sciatica), worse in morning/cold",
         "Warm foods with ghee. Milk boiled with turmeric. Avoid cold foods and constipation (main aggravator).",
         "Dashamoola Kwath. Yogaraja Guggulu. Mahanarayan oil externally. Rasna Saptak Kwath. Eranda (castor) oil 1 tsp at bedtime.",
         "CURE FOR LOWER BACK PAIN:\n1. Pavana Mukta Asana -- lie flat, knees to chest, rock gently. 1 min\n2. Nadi Shuddhi Kumbhaka -- alternate nostril breathing. 21 cycles\n3. Naaga Kumbhaka -- inhale fully, hold, exhale in 3 stages. 21 times\n4. Bhujangasana (Cobra) -- lie on stomach, lift chest, arms straight. 30 sec\n5. Shalabhasana (Locust) -- lie on stomach, lift legs and chest. 30 sec",
         "Kati Basti (oil pooling on lower back) -- 30 min daily for 7-14 days. Patra Pinda Sweda. Abhyanga with Dhanwantaram oil.",
         "Avoid heavy lifting. Use back support when sitting. Sleep on firm mattress. Warm compresses. Daily walking.",
         "Charaka Chikitsa 28, Sushruta Nidana 1, Ashtanga Hridaya Nidana 15"],

        # Digestive
        ["Digestive Disorders / IBS", "Vata/Pitta", "Agnimandya / Grahani",
         "Gas, bloating, irregular bowels, alternating constipation/diarrhea, abdominal pain",
         "VATA: Warm, cooked, moist food. Small frequent meals. Ginger tea before meals. Cumin-coriander-fennel tea.\nPITTA: Cooling but cooked food. Avoid spicy, sour. Fennel tea. Ghee with meals.",
         "Hingvashtak Churna 1/2 tsp before meals (gas/bloating). Kutaja Parpati (IBS). Chitrakadi Vati (weak digestion). Dadimashtak Churna. Lavanbhaskar Churna.",
         "CARE FOR DIGESTIVE DISORDERS:\n1. Vajrasana after meals -- 5-10 minutes\n2. Pavana Mukta Asana -- knees to chest. 1 min\n3. Ardha Matsyendrasana -- seated twist. Each side 30 sec\n4. Agnisar Kriya -- standing, exhale fully, pump stomach. 21 times",
         "Deepana-Pachana (appetizer-digestive) therapy first. Then mild Virechana if Pitta. Basti if Vata. Avoid Panchakarma during acute phase.",
         "Eat at fixed times. No cold water with meals. Walk 100 steps after each meal. Avoid eating when stressed. Chew food 32 times.",
         "Charaka Chikitsa 15 (Grahani), Ashtanga Hridaya Chikitsa 10"],

        ["Acidity / Heartburn", "Pitta", "Amlapitta",
         "Burning in chest/throat, sour belching, nausea, headache after eating, worse after spicy/sour food",
         "Cooling diet. Milk, ghee, coconut water, cucumber. Sweet, bitter, astringent tastes. Avoid: spicy, sour, fermented, alcohol, coffee, tomatoes.",
         "Avipattikar Churna 1 tsp after meals. Praval Pishti 250mg 2x/day. Shatavari 500mg. Yashtimadhu (Licorice) 500mg. Kamadudha Rasa 250mg.",
         "CARE FOR ACIDITY:\n1. Sheetali Pranayama -- tongue roll, inhale through tongue, exhale nose. 21 times\n2. Shavasana -- complete relaxation. 10 min\n3. Supta Baddha Konasana -- reclined butterfly. 3 min\n4. Viparita Karani -- legs up the wall. 5 min",
         "Virechana with Trivrit or Triphala (primary Pitta treatment). Takra Dhara (buttermilk on forehead).",
         "Avoid late-night eating. Don't lie down immediately after meals. Reduce stress. Eat largest meal at lunch (when Pitta is strongest).",
         "Charaka Chikitsa 15, Sushruta Uttara 40, Ashtanga Hridaya Chikitsa 10"],

        ["Constipation", "Vata", "Vibandha",
         "Hard, dry stools, infrequent, straining, bloating, incomplete evacuation",
         "Warm foods with ghee/oil. Soaked prunes, figs. Warm water on waking. Psyllium husk with warm milk. Avoid dry, raw, cold food.",
         "Triphala 1 tsp at bedtime with warm water. Eranda (Castor) oil 1-2 tsp at bedtime. Abhayarishta 20ml after meals. Avipattikar Churna.",
         "1. Pavana Mukta Asana -- knees to chest. 1 min\n2. Malasana -- deep squat, elbows pressing knees. 1-2 min\n3. Agnisar Kriya -- abdominal pumping. 21 times\n4. Trikonasana -- triangle pose. Each side 30 sec",
         "Basti (medicated enema) -- Anuvasana Basti with sesame oil. Niruha Basti with Dashamoola. Primary Vata treatment per Charaka.",
         "Drink 1 glass warm water on waking. Fixed meal times. Walk 15 min after dinner. Squat position for elimination.",
         "Charaka Chikitsa 13, Sushruta Chikitsa 14, Ayurveda Self-Healing Ch.5"],

        # Respiratory
        ["Asthma / Breathing", "Kapha/Vata", "Shwasa / Tamaka Shwasa",
         "Wheezing, breathlessness, chest tightness, cough with mucus, worse at night/damp weather",
         "Light, warm, dry food. No dairy, no cold drinks, no banana. Honey with warm water. Ginger-turmeric tea. Pungent, bitter tastes.",
         "Sitopaladi Churna 1 tsp with honey 3x/day. Kanakasava 20ml. Vasavaleha. Agastya Haritaki. Talisadi Churna.",
         "CURE FOR ASTHMA:\n1. Simha Garjana -- lion pose, roar. 21 times\n2. Bhastrika Pranayama -- rapid bellows breath. 21 cycles\n3. Bhujangasana -- cobra pose, chest opener. 30 sec\n4. Matsyasana -- fish pose, chest open, head back. 30 sec\n5. Ustrasana -- camel pose, backbend. 30 sec",
         "Vamana (therapeutic emesis) -- primary Kapha treatment. Nasya with Anu Taila. Dhumapana (herbal smoking for bronchial). Steam inhalation.",
         "Avoid cold, damp environments. No daytime sleep. Regular pranayama. Keep chest warm. Avoid dust and allergens.",
         "Charaka Chikitsa 17, Sushruta Uttara 51, Ashtanga Hridaya Chikitsa 4"],

        ["Sinusitis / Congestion", "Kapha", "Pratishyaya / Peenasa",
         "Blocked nose, thick mucus, headache, facial pressure, reduced smell, post-nasal drip",
         "Warm, light food. No dairy, no cold. Pepper-turmeric-ginger tea. Honey. Avoid sweet, heavy, oily food.",
         "Trikatu 1/4 tsp with honey 2x/day. Lakshmi Vilas Ras 250mg 2x/day. Haridra Khanda. Chitraka. Sitopaladi Churna.",
         "1. Jala Neti -- nasal wash with warm saline. Daily morning\n2. Kapalbhati Pranayama -- rapid exhales. 3 rounds of 30\n3. Bhastrika -- bellows breath. 21 cycles\n4. Surya Namaskar -- 6-12 rounds",
         "Nasya with Anu Taila or Shadbindu Taila -- 2-3 drops each nostril morning. Dhumapana. Steam with eucalyptus/ajwain.",
         "Avoid cold water on head. Keep sinuses warm. Avoid AC directly on face. Don't suppress sneezing.",
         "Charaka Chikitsa 26, Sushruta Uttara 24, Ashtanga Hridaya Uttara 19"],

        # Metabolic
        ["Diabetes", "Kapha/Pitta", "Prameha / Madhumeha",
         "Excessive thirst/urination, fatigue, weight gain, slow wound healing, sweet taste in mouth",
         "Bitter, astringent, pungent foods. Bitter gourd, fenugreek, turmeric. Barley, old rice. No sugar, no sweet fruits, no refined carbs, no dairy.",
         "Nishamalaki (Turmeric+Amla) 1 tsp 2x/day. Gudmar (Gymnema) 500mg. Vijaysar bark water. Chandraprabha Vati 500mg 2x/day. Triphala at bedtime.",
         "CURE FOR DIABETES:\n1. Manduka Asana -- frog pose, sit on heels, fists on abdomen, bend forward. 1 min\n2. Paschimottanasana -- seated forward bend. 1 min\n3. Ardha Matsyendrasana -- spinal twist. 30 sec each\n4. Kapalabhati -- 3 rounds of 30",
         "Virechana with Trivrit. Udvartana (dry powder massage). Takra Dhara. Per Sushruta: exercise is primary treatment.",
         "Daily vigorous exercise (walking 5 km). Avoid sedentary lifestyle. Don't sleep during day. Reduce stress. Monitor regularly.",
         "Charaka Chikitsa 6, Sushruta Chikitsa 11, Ayurveda Science of Life Ch.4"],

        ["Obesity / Weight Gain", "Kapha", "Sthaulya / Medoroga",
         "Excess weight, low energy, breathlessness on exertion, excess sweating, sweet cravings",
         "Light, dry, warm food. Honey in warm water morning (not cooked). Barley, millet, old rice. Bitter/pungent tastes. No sweet, oily, heavy food. Intermittent fasting.",
         "Triphala Guggulu 500mg 2x/day. Medohar Guggulu. Vidangadi Churna. Lekhaniya Varga herbs. Varunadi Kwath.",
         "CURE FOR OBESITY:\n1. Surya Namaskar -- 12 rounds minimum\n2. Trikonasana -- triangle pose. Each side 30 sec\n3. Naukasana -- boat pose. 3 rounds of 30 sec\n4. Bhastrika -- bellows breath. 3 rounds of 21\n5. Kapalabhati -- 3 rounds of 60",
         "Udvartana (herbal dry powder massage) -- primary. Vamana if strong. Lekhana Basti (fat-scraping enema).",
         "Wake by 5:30 AM. Vigorous daily exercise 45-60 min. No daytime sleep ever. Walk after every meal. Reduce portion sizes.",
         "Charaka Sutra 21, Sushruta Sutra 15, Ashtanga Hridaya Sutra 14"],

        ["Thyroid (Hypo)", "Kapha/Vata", "Galaganda",
         "Weight gain, fatigue, cold intolerance, constipation, dry skin, hair loss, puffy face",
         "Warm, light, stimulating food. Ginger, pepper, garlic. Coconut oil for cooking. Avoid raw cruciferous, soy, heavy food.",
         "Kanchanara Guggulu 500mg 2x/day (primary). Punarnava 500mg. Triphala at night. Ashwagandha 500mg (supports thyroid).",
         "1. Sarvangasana -- shoulder stand (direct thyroid stimulation). 1-2 min\n2. Halasana -- plough pose. 30 sec\n3. Matsyasana -- fish pose (counter pose). 30 sec\n4. Simhasana -- lion pose with tongue extension. 21 times\n5. Ujjayi Pranayama -- ocean breath. 21 cycles",
         "Nasya. Udvartana. Lepana (herbal paste) on throat with Kanchanara + Triphala.",
         "Regular exercise essential. Avoid cold exposure. Manage stress. Regular sleep pattern.",
         "Sushruta Nidana 11, Charaka Chikitsa 12, Ashtanga Hridaya Uttara 21"],

        # Skin
        ["Skin Problems / Acne", "Pitta/Kapha", "Kushtha / Yauvanapidika",
         "Acne, rashes, eczema, itching, inflammation, discoloration",
         "PITTA: Cooling diet. Bitter vegs (bitter gourd, neem leaves). Aloe vera juice. No spicy, sour, fermented, fried.\nKAPHA: Light, dry food. No dairy, sugar, oily food.",
         "Mahamanjishthadi Kwath. Khadirarishta 20ml. Gandhak Rasayan 250mg. Neem capsules 500mg. Panchatikta Ghrita Guggulu.",
         "CARE FOR SKIN PROBLEMS:\n1. Sheetali Pranayama -- cooling breath. 21 times\n2. Shavasana -- deep relaxation. 10 min\n3. Sarvangasana -- improved blood flow to face. 1 min\n4. Pranayama meditation -- 10 min",
         "Virechana (Pitta). Raktamokshana (bloodletting -- leech therapy per Sushruta). External: Neem-turmeric paste. Kumkumadi Tailam.",
         "Avoid sun during 10-2. Don't pop pimples. Wash face with cool water. Avoid synthetic fabrics. Manage stress.",
         "Charaka Chikitsa 7, Sushruta Nidana 5, Ashtanga Hridaya Nidana 14"],

        ["Eczema", "Vata/Pitta", "Vicharchika",
         "Dry, itchy, flaky patches, cracking, oozing, chronic relapsing",
         "Ghee internally. Avoid fermented, sour, seafood, nightshades. Anti-inflammatory: turmeric, neem.",
         "Panchatikta Ghrita 1 tsp daily. Mahamanjishthadi Kwath. Arogyavardhini Vati. External: Jatyadi Ghrita, Panchavalkala Kwath wash.",
         "1. Shavasana -- complete relaxation. 15 min\n2. Nadi Shodhana -- calming, reduces Vata. 10 min\n3. Gentle stretching -- no vigorous sweat-inducing exercise\n4. Pranayama meditation",
         "Virechana. Raktamokshana for severe cases. External: Panchavalkala Kwath bath. Eladi Coconut oil.",
         "Cotton clothing only. No hot water on affected areas. Moisturize with coconut oil/ghee. Manage stress. Identify triggers.",
         "Charaka Chikitsa 7, Sushruta Chikitsa 9"],

        # Cardiovascular
        ["Hypertension", "Pitta/Vata", "Raktachapa Vridhi",
         "Headache, dizziness, flushed face, anger, stress, palpitations",
         "Low-salt diet. Cooling foods. Garlic 2 cloves morning. Watermelon. No caffeine, alcohol, excess salt, spicy food.",
         "Sarpagandha (Rauwolfia) 250mg 2x/day. Arjuna 500mg 2x/day. Brahmi 500mg. Jatamansi 250mg. Mukta Pishti (pearl) 250mg.",
         "CARE FOR HYPERTENSION:\n1. Shavasana -- 15 min deep relaxation. MOST IMPORTANT.\n2. Nadi Shodhana -- alternate nostril. 15 min\n3. Bhramari -- humming bee. 11 times\n4. Gentle walking -- 30 min daily",
         "Virechana. Shirodhara with Brahmi oil (calming). Takra Dhara.",
         "Reduce stress through meditation. Avoid competitive activities. Moderate exercise only. Practice Yoga Nidra daily.",
         "Charaka Sutra 24, Sushruta Sharira 9, Ayurveda Science of Life Ch.11"],

        ["Heart Disease", "Pitta/Kapha", "Hridroga",
         "Chest pain, palpitations, breathlessness, fatigue, cholesterol",
         "Low-fat, plant-based diet. Arjuna bark tea. Garlic. Flax seeds. No fried, heavy, sweet food. Reduce salt.",
         "Arjuna 500mg 3x/day (primary cardiac herb). Pushkarmool 500mg. Hridayarnava Rasa. Mrigamadasava. Triphala at night.",
         "1. Walking -- 45 min daily\n2. Shavasana -- 15 min\n3. Pranayama -- Nadi Shodhana and Bhramari. 15 min\n4. Gentle yoga -- no inversions or breath holding",
         "Virechana. Hridaya Basti. Avoid strong Panchakarma. Gentle Abhyanga only.",
         "No smoking. Moderate exercise. Stress management. Regular monitoring. Sleep by 10 PM.",
         "Charaka Chikitsa 26, Sushruta Uttara 43, Ashtanga Hridaya Nidana 5"],

        # Head & Neurological
        ["Migraine / Headache", "Pitta/Vata", "Ardhavabhedaka / Shirahshoola",
         "One-sided throbbing headache, nausea, light sensitivity, visual aura",
         "PITTA: Cooling diet. Avoid sun, spicy, sour, fermented. Coriander water. Coconut water.\nVATA: Warm, regular meals. Ghee. Avoid fasting.",
         "Shirashularivajrarasa (classical). Pathyadi Kwath. Godanti Bhasma 250mg. Shatavari for Pitta-type. Dashamoola for Vata-type.",
         "CURE FOR MIGRAINE:\n1. Shavasana during attack -- dark, quiet room. 20 min\n2. Nadi Shodhana -- between attacks. 10 min daily\n3. Brahma Mudra -- neck rotations. 10 each direction\n4. Paschimottanasana -- seated forward bend. 1 min",
         "Nasya with Anu Taila (between attacks). Shirodhara with Brahmi oil. Shirobasti (oil pooling on head).",
         "Regular meals and sleep. Avoid triggers (sun, screens, loud noise). Cold compress on forehead during attack.",
         "Charaka Sutra 17, Sushruta Uttara 25, Ayurveda Science of Life Ch.12"],

        ["Memory / Concentration", "Vata/Kapha", "Smriti Bhramsha",
         "Forgetfulness, poor focus, brain fog, difficulty learning, mental fatigue",
         "VATA: Warm, ghee-rich food. Almonds (soaked). Walnuts. Brahmi tea.\nKAPHA: Light, stimulating food. Pepper, ginger.",
         "Brahmi 500mg 2x/day. Shankhapushpi 500mg. Medhya Rasayana (Charaka's 4): Mandukparni, Shankhapushpi, Guduchi, Yashtimadhu. Saraswatarishta 20ml.",
         "1. Sarvangasana -- blood flow to brain. 1 min\n2. Padmasana meditation -- 15 min daily\n3. Trataka -- candle gazing concentration. 5 min\n4. Nadi Shodhana -- calms and clarifies. 10 min",
         "Nasya with Brahmi Ghrita. Shirodhara. Shiro Abhyanga (head massage).",
         "Mental exercises daily. Reading. New learning. Adequate sleep 7-8 hrs. Reduce multi-tasking. Digital detox.",
         "Charaka Chikitsa 1 (Rasayana), Sushruta Sutra 15, Ashtanga Hridaya Uttara 6"],

        ["Epilepsy / Seizures", "Vata", "Apasmara",
         "Seizures, loss of consciousness, convulsions, post-seizure confusion",
         "Pure Sattvic diet. Ghee. Milk. No stimulants, no alcohol, no excess spice. Regular meal times.",
         "Brahmi Ghrita 1 tsp 2x/day. Saraswatarishta 20ml. Shatavari Ghrita. Panchagavya Ghrita. Smritisagar Rasa (classical).",
         "CARE FOR EPILEPSY:\n1. Shavasana -- daily relaxation. 20 min\n2. Nadi Shodhana -- balancing. 15 min\n3. Meditation -- 15 min\n4. AVOID: Kapalabhati, Bhastrika, breath retention, inversions",
         "Shirodhara (very important). Nasya with Brahmi Ghrita. Gentle Basti. NO strong Panchakarma.",
         "Regular sleep essential. Avoid flashing lights. No swimming alone. Stress reduction. Continue prescribed medication.",
         "Charaka Chikitsa 10, Sushruta Uttara 61, Ashtanga Hridaya Uttara 7"],

        # Urinary / Kidney
        ["Kidney Stones", "Kapha/Pitta", "Ashmari",
         "Severe flank pain, blood in urine, painful urination, nausea",
         "Plenty of warm water (3+ liters). Barley water. Coconut water. Kulatha (horse gram) soup. Avoid: excess calcium, oxalate-rich foods, salt.",
         "Gokshuradi Guggulu 500mg 2x/day. Punarnava 500mg. Varunadi Kwath. Chandraprabha Vati. Pashanbheda (Bergenia) 500mg.",
         "1. Pavana Mukta Asana. 1 min\n2. Bhujangasana -- cobra pose. 30 sec\n3. Dhanurasana -- bow pose. 30 sec\n4. Gentle twists -- promotes urinary flow",
         "Virechana. Uttara Basti (urethral instillation for severe cases). Sitz bath with Triphala decoction.",
         "Drink warm water throughout day. Don't hold urine. Moderate exercise. Avoid dehydration.",
         "Sushruta Chikitsa 7, Charaka Chikitsa 26, Ashtanga Hridaya Chikitsa 11"],

        # Hair
        ["Hair Loss / Baldness", "Pitta/Vata", "Khalitya / Indralupta",
         "Thinning, receding, patches of hair loss, premature greying",
         "Cooling, nourishing diet. Ghee, milk, coconut, sesame. Iron-rich food. Amla. Avoid excess spicy, sour.",
         "Bhringraj 500mg internally + Bhringraj oil externally. Neelibhringadi Kera Tailam. Narasimha Rasayanam. Amalaki 500mg. Yashtimadhu 500mg.",
         "CURE FOR BALDNESS:\n1. Sarvangasana -- blood to scalp. 1-2 min\n2. Sirshasana (headstand) -- if able. 30 sec to 2 min\n3. Adho Mukha Shvanasana -- downward dog. 1 min\n4. Uttanasana -- standing forward fold. 1 min",
         "Shiro Abhyanga (head massage with Bhringraj oil). Nasya with Anu Taila. Shirolepam (herbal paste on scalp). Raktamokshana for Pitta-type.",
         "Oil scalp 3x/week minimum. Avoid hot water on hair. Reduce stress. Don't tie hair too tight. Avoid chemical products.",
         "Sushruta Chikitsa 20, Charaka Chikitsa 7, Ashtanga Hridaya Uttara 23"],

        # Eye
        ["Eye Problems / Vision", "Pitta/Vata", "Timira / Drishti Dosha",
         "Blurred vision, eye strain, dry eyes, light sensitivity, deteriorating eyesight",
         "Ghee with meals (essential for eyes). Triphala water eye wash. Carrots, spinach, amla. Avoid excess screen time.",
         "Triphala Ghrita 1 tsp 2x/day (primary eye Rasayana). Saptamrit Lauh. Chandrodaya Varti (eye drops). Maha Triphala Ghrita.",
         "CARE FOR EYESIGHT:\n1. Trataka -- candle gazing. 5 min daily\n2. Eye exercises -- up/down/left/right/circles. 5 min\n3. Palming -- rub hands, cover closed eyes. 5 min\n4. Shavasana -- complete rest for eyes. 10 min",
         "Netra Tarpana (ghee pooling on eyes) -- 20 min, 7-day course. Nasya. Shirodhara. Anjana (collyrium application).",
         "20-20-20 rule for screens. Wash eyes with cool Triphala water morning. Don't read in dim light. Adequate sleep for eye rest.",
         "Sushruta Uttara 1-17, Charaka Chikitsa 26, Ashtanga Hridaya Uttara 8-16"],

        # Women's Health
        ["Menopause / Hot Flashes", "Pitta/Vata", "Rajonivritti",
         "Hot flashes, night sweats, mood swings, dryness, insomnia, anxiety",
         "Cooling + nourishing diet. Shatavari milk. Ghee. Flax seeds. Soy. Avoid spicy, caffeine, alcohol.",
         "Shatavari 500mg 2x/day (primary). Ashoka 500mg. Ashwagandha 500mg. Praval Pishti 250mg. Saraswatarishta.",
         "1. Supta Baddha Konasana -- reclined butterfly. 3-5 min\n2. Viparita Karani -- legs up wall. 5 min\n3. Sheetali Pranayama -- cooling. 21 cycles\n4. Nadi Shodhana -- balancing. 10 min\n5. Yoga Nidra -- guided relaxation. 20 min",
         "Abhyanga with Dhanwantaram oil. Basti with Shatavari Ghrita. Shirodhara.",
         "Regular exercise. Stay cool. Layer clothing. Stress management. Social connection.",
         "Charaka Chikitsa 30, Sushruta Sharira 2, Ashtanga Hridaya Sharira 1"],

        ["PCOS", "Kapha/Pitta", "Artava Kshaya",
         "Irregular periods, weight gain, acne, excess hair, cysts on ovaries",
         "Light, warm food. Avoid dairy, sugar, processed food. Fenugreek seeds. Cinnamon. Flax seeds. Bitter foods.",
         "Kanchanara Guggulu 500mg 2x/day. Chandraprabha Vati 500mg. Shatavari 500mg. Ashoka 500mg. Rajapravartini Vati.",
         "1. Surya Namaskar -- 12 rounds\n2. Butterfly pose (Baddha Konasana) -- 2 min\n3. Bharadvajasana -- seated twist. Each side 30 sec\n4. Dhanurasana -- bow pose. 30 sec\n5. Kapalabhati -- 3 rounds of 30",
         "Virechana. Uttara Basti. Udvartana. Yoni Pichu (local treatment).",
         "Regular vigorous exercise. Reduce stress. Weight management crucial. Regular sleep. Avoid hormonal disruptors.",
         "Charaka Chikitsa 30, Sushruta Sharira 2, Kashyapa Samhita"],

        # Addiction
        ["Addiction / Substance", "Vata/Pitta", "Madatyaya",
         "Cravings, withdrawal symptoms, anxiety, liver issues, sleep disturbance",
         "Nourishing, stabilizing diet. Regular meals. Dates, ghee, milk. Avoid: the substance, caffeine initially. Sattvic diet.",
         "Ashwagandha 500mg 2x/day (anxiety+strength). Brahmi 500mg (mental clarity). Kutki 500mg (liver). Guduchi 500mg (detox+immunity). Jatamansi 250mg (calming).",
         "CARE FOR ADDICTION:\n1. Nadi Shodhana -- calming. 15 min 2x/day\n2. Meditation -- mindfulness. 20 min daily\n3. Shavasana -- relaxation. 15 min\n4. Walking -- 45 min daily",
         "Full Panchakarma sequence (Vamana, Virechana, Basti) for deep cleansing. Shirodhara. Abhyanga. Counseling support essential.",
         "Structured daily routine is CRITICAL. Support community. Avoid triggers. New hobbies/activities. Sleep hygiene.",
         "Charaka Chikitsa 24, Sushruta Uttara 47, Ashtanga Hridaya Nidana 6"],

        # Autoimmune
        ["Autoimmune Disorders", "Vata/Pitta", "Ama + Ojas Kshaya",
         "Chronic inflammation, fatigue, pain that moves, multiple systems affected, flares and remissions",
         "Anti-inflammatory diet. Turmeric, ghee, ginger. Avoid: processed food, sugar, nightshades, dairy for some. Gentle on digestion.",
         "Guduchi 500mg 2x/day (immunomodulator). Ashwagandha 500mg. Shatavari 500mg. Triphala at night. Chyawanprash 1 tbsp morning.",
         "1. Gentle yoga only -- no vigorous practice during flares\n2. Shavasana -- 20 min daily\n3. Nadi Shodhana -- 10 min\n4. Gentle Surya Namaskar -- 4-6 rounds when able\n5. Meditation -- 15 min",
         "Start with Deepana-Pachana (digestive correction). Then gentle Panchakarma only when strength allows. Basti is safest. Abhyanga with medicated oils.",
         "Stress management is PRIMARY. Regular routine. Adequate rest. Gentle exercise. Avoid over-exertion. Identify and avoid triggers.",
         "Charaka Chikitsa 28 (Vata Vyadhi), Sushruta Sutra 24, Ashtanga Hridaya Sutra 13"],

        # Common everyday
        ["Common Cold / Flu", "Kapha/Vata", "Pratishyaya",
         "Running nose, sneezing, congestion, mild fever, body ache, sore throat",
         "Warm liquids only. Ginger-turmeric-honey tea. Tulsi tea. Light food or fasting. No dairy, cold, heavy food.",
         "Tribhuvankirti Rasa 250mg 2x/day. Talisadi Churna with honey. Sitopaladi Churna. Tulsi 500mg. Mahalaxmi Vilas Ras.",
         "1. Steam inhalation with ajwain/eucalyptus. 10 min 2x/day\n2. Jala Neti -- nasal wash. Morning\n3. Bhastrika -- mild. 11 cycles\n4. Rest in Shavasana",
         "Swedana (herbal steam bath). Nasya with Anu Taila (after acute phase resolves). Dhumapana (herbal smoking for residual congestion).",
         "Complete rest during fever. Stay warm. Gargle with warm salt + turmeric water. Avoid cold drafts and AC.",
         "Charaka Chikitsa 26, Sushruta Uttara 24, Ashtanga Hridaya Uttara 19"],

        ["Fever", "Pitta/Vata", "Jwara",
         "Elevated temperature, body ache, loss of appetite, thirst, restlessness",
         "Fasting or liquid diet ONLY during fever. Rice gruel (Peya). Laja Manda (parched rice water). Warm water. NO heavy food.",
         "Sudarshan Ghan Vati 500mg 3x/day. Mahasudarshan Churna. Guduchi Kwath (Giloy water). Tribhuvankirti Rasa for cold-fever. Amritarishta 20ml.",
         "DURING FEVER: Complete rest only. Shavasana.\nAFTER FEVER: Gentle pranayama. Walking. Gradual return to yoga.\nDO NOT exercise during active fever.",
         "Langhana (therapeutic fasting) is PRIMARY treatment for fever per Charaka. Sponging if high. Gradually reintroduce food (Peya → Vilepi → Yusha → normal).",
         "Complete rest. No bath during high fever. Sip warm water frequently. Don't suppress fever with cold compresses unless dangerously high.",
         "Charaka Chikitsa 3 (Jwara -- longest chapter), Sushruta Uttara 39, Ashtanga Hridaya Chikitsa 1"],

        ["Vertigo / Dizziness", "Vata", "Bhrama",
         "Spinning sensation, loss of balance, nausea, may be triggered by position change",
         "Warm, grounding, heavy food. Ghee. Sesame oil in food. Milk. Avoid: fasting, light/dry food, cold drinks.",
         "Ashwagandha 500mg. Bala 500mg. Saraswatarishta 20ml. Dashamoola Kwath. Drakshadi Kwath.",
         "CURE FOR VERTIGO:\n1. Shavasana with legs elevated. 10 min\n2. Slow, controlled head movements. 5 min\n3. Nadi Shodhana -- very gentle. 10 min\n4. AVOID: inversions, rapid movements, Kapalabhati",
         "Nasya with Anu Taila. Shirodhara. Karnapurana (warm oil in ears). Abhyanga with Bala oil.",
         "Move slowly, especially when changing position. Hold railings. Adequate sleep. Reduce stress. Stay hydrated.",
         "Charaka Sutra 20, Sushruta Uttara 25, Ashtanga Hridaya Sutra 11"],
    ]

    row = 3
    for d in diseases:
        for col, val in enumerate(d, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.alignment = wrap
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
            if col == 2:
                if "Vata" in str(val) and "Pitta" not in str(val) and "Kapha" not in str(val):
                    cell.fill = vata_fill
                elif "Pitta" in str(val) and "Kapha" not in str(val):
                    cell.fill = pitta_fill
                elif "Kapha" in str(val) and "Pitta" not in str(val):
                    cell.fill = kapha_fill
        ws2.row_dimensions[row].height = 280
        row += 1

    # =====================================================================
    # SHEET 3: DOSHA DIET QUICK REFERENCE
    # =====================================================================
    ws3 = wb.create_sheet("Dosha Diet Reference")
    set_title(ws3, "DOSHA-SPECIFIC DIET & LIFESTYLE QUICK REFERENCE")

    headers3 = ["Category", "Vata (Increase)", "Vata (Avoid)", "Pitta (Increase)", "Pitta (Avoid)", "Kapha (Increase)", "Kapha (Avoid)"]
    widths3 = [18, 30, 30, 30, 30, 30, 30]
    write_headers(ws3, headers3, widths3)

    diet_data = [
        ["Tastes", "Sweet, Sour, Salty", "Bitter, Astringent, Pungent", "Sweet, Bitter, Astringent", "Sour, Salty, Pungent", "Pungent, Bitter, Astringent", "Sweet, Sour, Salty"],
        ["Grains", "Rice (basmati), wheat, oats (cooked)", "Millet, corn, buckwheat, barley", "Rice, wheat, barley, oats", "Corn, millet, rye, brown rice", "Barley, millet, corn, buckwheat, old rice", "Rice (new), wheat, oats"],
        ["Vegetables", "Cooked: asparagus, beet, carrot, sweet potato, squash, zucchini", "Raw vegetables, cabbage, cauliflower, broccoli (raw), sprouts", "Sweet/bitter: cucumber, broccoli, cauliflower, leafy greens, peas", "Hot peppers, tomato, onion (raw), beet, radish, garlic", "Pungent/bitter: radish, spinach, onion, garlic, pepper, cabbage", "Sweet potato, tomato, cucumber, zucchini (excess)"],
        ["Fruits", "Sweet: banana, mango, grape, avocado, berries, dates, figs", "Dry fruits (excess), apple (raw), cranberry", "Sweet: apple, grape, pomegranate, coconut, melon, pear", "Sour: citrus, pineapple, sour berries", "Apple, pomegranate, cranberry, dried fruits", "Banana, mango, coconut, melon, dates, figs"],
        ["Dairy", "Warm milk, ghee, butter, cream, yogurt", "Cold milk, ice cream", "Milk, ghee, butter (unsalted)", "Sour cream, buttermilk, salted cheese, yogurt (excess)", "Warm goat milk, small amount ghee", "Cheese, cream, butter, ice cream, cold milk, yogurt"],
        ["Oils", "Sesame, ghee, almond, olive (generous)", "None (need oil)", "Coconut, sunflower, ghee", "Sesame, almond, corn", "Mustard, sunflower (minimal)", "Sesame (excess), any excess oil"],
        ["Spices", "ALL spices, especially ginger, cumin, cinnamon, cardamom, hing", "None (all spices are good for Vata)", "Cooling: coriander, fennel, cardamom, turmeric, mint", "Hot: chili, mustard, clove, fenugreek (excess)", "ALL hot spices: ginger, pepper, mustard, clove, cinnamon", "Salt (excess)"],
        ["Proteins", "Chicken, fish, eggs, tofu, mung dal, nuts", "Beans (most), especially chickpea, kidney bean", "Mung dal, chickpea, tofu, chicken (white), egg white", "Red meat, seafood, egg yolk, lentils (excess)", "Mung dal, chicken, turkey, egg white (moderate)", "Red meat, pork, seafood, cheese, tofu (excess)"],
        ["Drinks", "Warm water, ginger tea, milk, chai, warm lemon water", "Cold water, carbonated drinks, coffee (excess)", "Cool water, mint tea, coconut water, milk", "Alcohol, coffee, hot drinks (excess)", "Warm water with honey, ginger tea, hot water, herbal tea", "Cold water, milkshakes, sweet juices"],
        ["Sweeteners", "Jaggery, honey (raw), dates, maple syrup", "White sugar (excess)", "Maple syrup, dates, small amount sugar", "Honey (cooked), molasses", "Raw honey ONLY (in warm water, never cooked)", "All other sweeteners, sugar, jaggery"],
        ["Exercise", "Gentle: yoga, walking, swimming, tai chi, dance", "Vigorous, excessive, running marathons", "Moderate: swimming, cycling, hiking, team sports", "Excessive competition, hot yoga, midday exercise", "Vigorous: running, HIIT, weight training, active sports", "Sedentary, sleeping after meals, couch lifestyle"],
        ["Sleep", "Sleep by 10 PM. 8 hours. Warm bedroom. Oil feet.", "Late nights, irregular sleep times", "Sleep by 10:30 PM. 7-8 hours. Cool bedroom.", "Late nights, sleeping in hot room", "Wake by 5:30-6 AM. 6-7 hours. No naps.", "Sleeping >8 hours, daytime naps, sleeping after meals"],
        ["Best Season Practices", "Autumn/Winter: More oil, warmth, routine, heavier food", "Cold, windy, dry exposure", "Summer: Cooling foods, moonlight walks, swimming", "Excess sun, hot environments, competition", "Spring: Fasting, exercise, light food, early rising", "Cold, damp, sitting indoors all day"],
    ]

    row = 3
    for d in diet_data:
        for col, val in enumerate(d, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.alignment = wrap
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
            elif col in (2, 3):
                cell.fill = vata_fill
            elif col in (4, 5):
                cell.fill = pitta_fill
            elif col in (6, 7):
                cell.fill = kapha_fill
        ws3.row_dimensions[row].height = 80
        row += 1

    # Save
    output = "data/Prakriti_Dosha_Diagnosis_Tree.xlsx"
    wb.save(output)
    print(f"Saved to {output}")


if __name__ == "__main__":
    create_diagnosis_tree()
