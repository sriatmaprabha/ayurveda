"""Extract ALL knowledge into structured JSONL chunks in data1/ folder.

Categories:
1. dosha_knowledge.jsonl - All dosha-related content (Vata, Pitta, Kapha characteristics, imbalances)
2. herb_knowledge.jsonl - All herb mentions, formulations, preparations, dosages
3. food_diet_knowledge.jsonl - All dietary recommendations, food properties, Rasa/Virya/Vipaka
4. asana_knowledge.jsonl - All yoga asanas with techniques, benefits, contraindications
5. disease_treatment.jsonl - All disease descriptions, symptoms, treatments from texts
6. panchakarma_therapy.jsonl - All Panchakarma and therapeutic procedures
7. anatomy_physiology.jsonl - Dhatu, Srotas, Agni, Ojas, organ systems
8. diagnosis_methods.jsonl - Pariksha methods, pulse, tongue, nail diagnosis
9. formulations.jsonl - All compound formulations (Churna, Kwath, Arishta, Vati, Ghrita, Bhasma)
10. scripture_verses.jsonl - Direct verses/quotes from classical texts
11. daily_seasonal_routine.jsonl - Dinacharya, Ritucharya, lifestyle
12. mental_health.jsonl - Manas Prakriti, psychology, meditation, spiritual health
13. master_index.jsonl - Complete combined dataset for RAG

Each record: {"id", "category", "subcategory", "text", "source_file", "source_book", "page_or_section", "dosha_relevance", "condition_relevance", "tags"}
"""

import json
import logging
import re
import sys
import csv
from pathlib import Path
from collections import defaultdict

csv.field_size_limit(sys.maxsize)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

PROJECT_ROOT = Path(__file__).parent
DATA1 = PROJECT_ROOT / "data1"
DATA1.mkdir(exist_ok=True)

# =====================================================================
# CATEGORY KEYWORDS for classification
# =====================================================================
CATEGORY_KEYWORDS = {
    "dosha_knowledge": [
        "vata", "pitta", "kapha", "tridosha", "dosha", "prakriti", "vikriti",
        "constitution", "body type", "temperament", "guna", "sattva", "rajas", "tamas",
    ],
    "herb_knowledge": [
        "herb", "plant", "drug", "medicine", "aushadha", "dravya", "botanical",
        "ashwagandha", "brahmi", "triphala", "guggulu", "neem", "tulsi", "shatavari",
        "haritaki", "amalaki", "bibhitaki", "pippali", "guduchi", "turmeric", "ginger",
        "churna", "kwath", "decoction", "extract", "powder", "formulation",
    ],
    "food_diet_knowledge": [
        "diet", "food", "ahara", "rasa", "virya", "vipaka", "taste", "sweet", "sour",
        "salty", "pungent", "bitter", "astringent", "milk", "ghee", "honey",
        "grain", "fruit", "vegetable", "meal", "fasting", "nutrition",
    ],
    "asana_knowledge": [
        "asana", "yoga", "posture", "pose", "pranayama", "kumbhaka", "mudra",
        "bandha", "surya namaskar", "meditation", "dhyana", "trataka",
        "kapalabhati", "bhastrika", "nadi shodhana",
    ],
    "disease_treatment": [
        "disease", "roga", "chikitsa", "treatment", "cure", "therapy", "symptom",
        "nidana", "cause", "pathogenesis", "samprapti", "prognosis", "sadhya",
        "fever", "jwara", "cough", "kasa", "asthma", "shwasa", "diabetes", "prameha",
        "arthritis", "kushtha", "skin", "pandu", "anemia",
    ],
    "panchakarma_therapy": [
        "panchakarma", "vamana", "virechana", "basti", "nasya", "raktamokshana",
        "purification", "detox", "shodhana", "shamana", "oleation", "snehana",
        "swedana", "fomentation", "enema", "purgation", "emesis",
    ],
    "anatomy_physiology": [
        "dhatu", "rasa", "rakta", "mamsa", "meda", "asthi", "majja", "shukra",
        "srotas", "channel", "agni", "ojas", "tejas", "prana", "organ", "tissue",
        "marma", "nadi", "sharira", "anatomy", "physiology",
    ],
    "diagnosis_methods": [
        "pariksha", "diagnosis", "examination", "pulse", "nadi", "tongue", "jihva",
        "eye", "netra", "nail", "nakha", "urine", "mutra", "stool", "mala",
        "ashtavidha", "dashavidha", "trividha",
    ],
    "formulations": [
        "formulation", "preparation", "yoga", "compound", "bhasma", "rasa",
        "arishta", "asava", "avaleha", "lehya", "ghrita", "taila", "vati",
        "gutika", "mandura", "pishti", "lauh", "guggulu",
    ],
    "daily_seasonal_routine": [
        "dinacharya", "ritucharya", "routine", "daily", "seasonal", "morning",
        "sleep", "exercise", "bath", "oil massage", "abhyanga", "lifestyle",
        "svastha", "preventive", "health preservation",
    ],
    "mental_health": [
        "mental", "mind", "manas", "anxiety", "depression", "stress", "fear",
        "anger", "meditation", "yoga nidra", "concentration", "memory", "medhya",
        "unmada", "apasmara", "insomnia", "anidra", "consciousness",
    ],
}


def classify_text(text: str) -> list[str]:
    """Classify text into one or more categories based on keywords."""
    text_lower = text.lower()
    categories = []
    scores = {}

    for cat, keywords in CATEGORY_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score >= 2:
            scores[cat] = score

    if scores:
        # Return top 3 categories by score
        sorted_cats = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        categories = [cat for cat, _ in sorted_cats[:3]]
    else:
        categories = ["general"]

    return categories


def extract_dosha_relevance(text: str) -> list[str]:
    """Identify which doshas are mentioned."""
    text_lower = text.lower()
    doshas = []
    if "vata" in text_lower or "vayu" in text_lower:
        doshas.append("Vata")
    if "pitta" in text_lower:
        doshas.append("Pitta")
    if "kapha" in text_lower or "shleshma" in text_lower:
        doshas.append("Kapha")
    return doshas


def extract_condition_tags(text: str) -> list[str]:
    """Extract disease/condition mentions."""
    text_lower = text.lower()
    conditions = []
    condition_keywords = {
        "fever": "Jwara", "cough": "Kasa", "asthma": "Shwasa",
        "diabetes": "Prameha", "anemia": "Pandu", "skin disease": "Kushtha",
        "arthritis": "Sandhivata", "diarrhea": "Atisara", "constipation": "Vibandha",
        "headache": "Shirahshoola", "eye disease": "Netra Roga",
        "heart disease": "Hridroga", "obesity": "Sthaulya",
        "piles": "Arsha", "edema": "Shotha", "wound": "Vrana",
        "insomnia": "Anidra", "anxiety": "Chittodvega",
        "depression": "Vishada", "epilepsy": "Apasmara",
    }
    for keyword, ayurvedic_name in condition_keywords.items():
        if keyword in text_lower:
            conditions.append(ayurvedic_name)
    return conditions


def get_book_name(file_name: str) -> str:
    """Extract readable book name from file name."""
    name = file_name.replace(".pdf", "").replace(".csv", "").replace(".md", "")
    name = name.replace(".txt", "").replace(".oneocr", "").replace(".sql", "")
    # Clean up common patterns
    name = re.sub(r"^\d{4}\.\d+\.", "", name)  # Remove year.number prefix
    name = name.replace("-", " ").replace("_", " ").strip()
    return name[:100]


def process_existing_chunks():
    """Load and reclassify all existing processed chunks."""
    chunk_files = [
        PROJECT_ROOT / "data" / "processed" / "chunks.jsonl",
        PROJECT_ROOT / "data" / "processed" / "ngpt_chunks.jsonl",
        PROJECT_ROOT / "data" / "processed" / "ocr_chunks.jsonl",
    ]

    all_records = []
    for chunk_file in chunk_files:
        if not chunk_file.exists():
            continue
        logging.info(f"Loading {chunk_file.name}...")
        with open(chunk_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        all_records.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue

    logging.info(f"Loaded {len(all_records)} total chunks from processed files")
    return all_records


def process_excel_data():
    """Extract structured data from Excel sheets."""
    records = []

    # Load Vikriti data
    try:
        from openpyxl import load_workbook
        vf = PROJECT_ROOT / "data" / "Vikriti_Dosha_Diagnosis_Guide.xlsx"
        if vf.exists():
            wb = load_workbook(vf)
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=3, values_only=False):
                    values = [cell.value for cell in row if cell.value]
                    if values:
                        text = " | ".join(str(v)[:500] for v in values if v)
                        if len(text) > 50:
                            records.append({
                                "text": text,
                                "source_file": "Vikriti_Dosha_Diagnosis_Guide.xlsx",
                                "section_title": ws.title,
                                "section_type": "vikriti_guide",
                            })
            logging.info(f"Extracted {len(records)} records from Vikriti Excel")

        # Load Prakriti data
        pf = PROJECT_ROOT / "data" / "Prakriti_Dosha_Diagnosis_Tree_v2.xlsx"
        if pf.exists():
            wb = load_workbook(pf)
            count = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=3, values_only=False):
                    values = [cell.value for cell in row if cell.value]
                    if values:
                        text = " | ".join(str(v)[:500] for v in values if v)
                        if len(text) > 50:
                            records.append({
                                "text": text,
                                "source_file": "Prakriti_Dosha_Diagnosis_Tree_v2.xlsx",
                                "section_title": ws.title,
                                "section_type": "prakriti_guide",
                            })
                            count += 1
            logging.info(f"Extracted {count} records from Prakriti Excel")
    except Exception as e:
        logging.warning(f"Excel extraction error: {e}")

    return records


def process_json_data():
    """Extract structured data from JSON files."""
    records = []

    # Asana protocols
    ap = PROJECT_ROOT / "data" / "asana_protocols.json"
    if ap.exists():
        with open(ap, "r", encoding="utf-8") as f:
            protocols = json.load(f)
        for name, data in protocols.items():
            if name == "Name of kriya":
                continue
            text = f"Protocol: {name}\nSequence: {data.get('steps_summary', '')}\n"
            for i, step in enumerate(data.get("step_details", []), 1):
                text += f"\nStep {i}:\n{step}\n"
            records.append({
                "text": text,
                "source_file": "asana_protocols.json",
                "section_title": name,
                "section_type": "asana_protocol",
            })
        logging.info(f"Extracted {len(records)} asana protocol records")

    # Asana full details
    ad = PROJECT_ROOT / "data" / "asana_full_details.json"
    if ad.exists():
        with open(ad, "r", encoding="utf-8") as f:
            asanas = json.load(f)
        count = 0
        for name, data in asanas.items():
            text = (
                f"Asana: {data.get('english_name', name)}\n"
                f"Sanskrit: {data.get('sanskrit_name', '')}\n"
                f"Description: {data.get('description', '')}\n"
                f"Difficulty: {data.get('difficulty', '')}\n"
                f"Technique: {data.get('technique', '')}\n"
                f"Benefits: {data.get('benefits', '')}\n"
                f"Tags: {data.get('goal_tags', '')}"
            )
            records.append({
                "text": text,
                "source_file": "asana_full_details.json",
                "section_title": name,
                "section_type": "asana_detail",
            })
            count += 1
        logging.info(f"Extracted {count} asana detail records")

    # Herb references
    hr = PROJECT_ROOT / "data" / "herb_references.json"
    if hr.exists():
        with open(hr, "r", encoding="utf-8") as f:
            herbs = json.load(f)
        count = 0
        for condition, refs in herbs.items():
            for ref in refs:
                text = f"Condition: {condition}\nSource: {ref.get('source', '')}\nSection: {ref.get('section', '')}\n\n{ref.get('text', '')}"
                records.append({
                    "text": text,
                    "source_file": "herb_references.json",
                    "section_title": f"{condition} - {ref.get('section', '')}",
                    "section_type": "herb_reference",
                })
                count += 1
        logging.info(f"Extracted {count} herb reference records")

    return records


def main():
    logging.info("=" * 60)
    logging.info("COMPREHENSIVE DATA EXTRACTION - Every book, every page")
    logging.info("=" * 60)

    # Step 1: Load all existing processed chunks
    all_records = process_existing_chunks()

    # Step 2: Add Excel data
    excel_records = process_excel_data()
    all_records.extend(excel_records)

    # Step 3: Add JSON structured data
    json_records = process_json_data()
    all_records.extend(json_records)

    logging.info(f"\nTotal records to process: {len(all_records)}")

    # Step 4: Classify every record and write to category-specific JSONL files
    category_files = {}
    category_counts = defaultdict(int)
    master_records = []
    record_id = 0

    for record in all_records:
        text = record.get("text", "")
        if not text or len(text.strip()) < 30:
            continue

        source_file = record.get("source_file", record.get("metadata", {}).get("file_name", "unknown"))
        section_title = record.get("section_title", record.get("metadata", {}).get("section_title", ""))
        section_type = record.get("section_type", record.get("metadata", {}).get("section_type", ""))

        categories = classify_text(text)
        doshas = extract_dosha_relevance(text)
        conditions = extract_condition_tags(text)

        record_id += 1
        structured = {
            "id": f"rec_{record_id:06d}",
            "category": categories[0],
            "subcategories": categories[1:] if len(categories) > 1 else [],
            "text": text,
            "source_file": source_file,
            "source_book": get_book_name(source_file) if source_file else "",
            "page_or_section": section_title,
            "section_type": section_type,
            "dosha_relevance": doshas,
            "condition_relevance": conditions,
            "tags": categories + doshas + conditions,
        }

        # Write to category file
        primary_cat = categories[0]
        if primary_cat not in category_files:
            fp = DATA1 / f"{primary_cat}.jsonl"
            category_files[primary_cat] = open(fp, "w", encoding="utf-8")
        category_files[primary_cat].write(json.dumps(structured, ensure_ascii=False) + "\n")
        category_counts[primary_cat] += 1

        # Also write to secondary category files
        for sec_cat in categories[1:]:
            if sec_cat not in category_files:
                fp = DATA1 / f"{sec_cat}.jsonl"
                category_files[sec_cat] = open(fp, "w", encoding="utf-8")
            category_files[sec_cat].write(json.dumps(structured, ensure_ascii=False) + "\n")
            category_counts[sec_cat] += 1

        master_records.append(structured)

    # Close all files
    for f in category_files.values():
        f.close()

    # Write master index
    logging.info("Writing master_index.jsonl...")
    with open(DATA1 / "master_index.jsonl", "w", encoding="utf-8") as f:
        for rec in master_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Print summary
    logging.info("\n" + "=" * 60)
    logging.info("EXTRACTION COMPLETE")
    logging.info("=" * 60)
    logging.info(f"Total records: {len(master_records)}")
    logging.info(f"Output directory: {DATA1}")
    logging.info("\nCategory breakdown:")
    for cat, count in sorted(category_counts.items(), key=lambda x: x[1], reverse=True):
        logging.info(f"  {cat:35s}: {count:6d} records")

    # Write summary
    summary = {
        "total_records": len(master_records),
        "total_source_files": len(set(r["source_file"] for r in master_records)),
        "category_counts": dict(sorted(category_counts.items(), key=lambda x: x[1], reverse=True)),
        "dosha_coverage": {
            "vata_mentions": sum(1 for r in master_records if "Vata" in r["dosha_relevance"]),
            "pitta_mentions": sum(1 for r in master_records if "Pitta" in r["dosha_relevance"]),
            "kapha_mentions": sum(1 for r in master_records if "Kapha" in r["dosha_relevance"]),
        },
        "condition_coverage": defaultdict(int),
    }
    for r in master_records:
        for c in r["condition_relevance"]:
            summary["condition_coverage"][c] += 1
    summary["condition_coverage"] = dict(sorted(
        summary["condition_coverage"].items(), key=lambda x: x[1], reverse=True
    ))

    with open(DATA1 / "extraction_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    logging.info(f"\nSummary saved to {DATA1 / 'extraction_summary.json'}")


if __name__ == "__main__":
    main()
