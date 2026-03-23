"""Generate Vikriti Dosha diagnostic and cure Excel sheet."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_vikriti_sheet():
    wb = Workbook()

    # ==========================================
    # SHEET 1: Single Dosha Vikriti
    # ==========================================
    ws1 = wb.active
    ws1.title = "Single Dosha Vikriti"

    single_dosha_data = [
        {
            "dosha": "Vata Vikriti",
            "element": "Air + Ether (Vayu + Akasha)",
            "body_type": "Thin frame, dry skin, cold hands/feet, light bones",
            "key_symptoms": (
                "1. Anxiety, restlessness, fear\n"
                "2. Dry, rough, cracked skin\n"
                "3. Constipation, bloating, gas\n"
                "4. Joint pain, cracking joints\n"
                "5. Insomnia, light/disturbed sleep\n"
                "6. Irregular appetite, weight loss\n"
                "7. Tremors, twitching, numbness\n"
                "8. Cold extremities"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Warm, moist, oily foods (ghee, sesame oil)\n"
                "- Sweet, sour, salty tastes\n"
                "- Warm milk with ashwagandha at night\n"
                "- Avoid raw, cold, dry foods\n\n"
                "HERBS:\n"
                "- Ashwagandha (strength, calming)\n"
                "- Bala (nourishing, grounding)\n"
                "- Dashamoola (10-root decoction for Vata)\n"
                "- Triphala (mild, for constipation)\n\n"
                "THERAPIES:\n"
                "- Abhyanga (warm sesame oil massage)\n"
                "- Basti (medicated enema - primary Vata treatment per Charaka)\n"
                "- Shirodhara (warm oil on forehead)\n\n"
                "ASANAS:\n"
                "- Padmasana (grounding)\n"
                "- Paschimottanasana (calming)\n"
                "- Shavasana (deep rest)\n\n"
                "LIFESTYLE:\n"
                "- Fixed daily routine (dinacharya)\n"
                "- Early sleep (before 10 PM)\n"
                "- Warm baths, avoid excessive travel"
            ),
            "followup_questions": (
                "1. How is your sleep? Do you wake up between 2-6 AM? (Vata time)\n"
                "2. Do you experience gas, bloating, or irregular bowel movements?\n"
                "3. Are your joints cracking or painful, especially in cold weather?\n"
                "4. Do you feel anxious or fearful without clear reason?\n"
                "5. Is your skin dry, rough, or flaking?\n"
                "6. Do you feel cold easily, especially hands and feet?\n"
                "7. Is your appetite irregular — sometimes hungry, sometimes not?\n"
                "8. Do you have any tremors, twitching, or restless legs?\n"
                "9. How is your menstrual cycle? (if applicable) — irregular/scanty?\n"
                "10. Do symptoms worsen in autumn/early winter or windy weather?"
            ),
            "aggravating_factors": "Autumn/early winter, cold/dry/windy weather, excess travel, fasting, late nights, raw food, bitter/astringent tastes",
            "source": "Ashtanga Hridaya Sutrasthana, Charaka Samhita Vol 2, Sushruta Samhita Vol 1",
        },
        {
            "dosha": "Pitta Vikriti",
            "element": "Fire + Water (Agni + Jala)",
            "body_type": "Medium frame, warm skin, sharp features, moderate weight",
            "key_symptoms": (
                "1. Anger, irritability, impatience\n"
                "2. Acid reflux, heartburn, ulcers\n"
                "3. Skin rashes, acne, inflammation\n"
                "4. Excessive sweating, body odor\n"
                "5. Burning sensation (eyes, stomach, skin)\n"
                "6. Diarrhea, loose stools\n"
                "7. Premature graying, hair loss\n"
                "8. Bloodshot or sensitive eyes"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Cooling foods (coconut, cucumber, melon)\n"
                "- Sweet, bitter, astringent tastes\n"
                "- Milk, ghee, cooling herbs\n"
                "- Avoid spicy, sour, fermented, oily foods\n\n"
                "HERBS:\n"
                "- Shatavari (cooling, nourishing)\n"
                "- Amalaki (Pitta-specific from Triphala)\n"
                "- Guduchi/Giloy (anti-inflammatory)\n"
                "- Neem (blood purification)\n"
                "- Chandana/Sandalwood (cooling)\n\n"
                "THERAPIES:\n"
                "- Virechana (therapeutic purgation - primary Pitta treatment per Charaka)\n"
                "- Cooling Abhyanga (coconut oil massage)\n"
                "- Netra Tarpana (ghee eye treatment)\n\n"
                "ASANAS:\n"
                "- Sheetali Pranayama (cooling breath)\n"
                "- Chandrasana (moon pose)\n"
                "- Shavasana (relaxation)\n\n"
                "LIFESTYLE:\n"
                "- Avoid midday sun\n"
                "- Moonlight walks\n"
                "- Cool environment, swimming\n"
                "- Moderate exercise, avoid competition"
            ),
            "followup_questions": (
                "1. Do you experience heartburn, acid reflux, or stomach burning?\n"
                "2. Are you prone to anger, frustration, or critical judgment?\n"
                "3. Do you have skin rashes, acne, or inflammatory conditions?\n"
                "4. Do you sweat excessively, even in mild weather?\n"
                "5. Are your eyes often red, burning, or sensitive to light?\n"
                "6. Do you experience loose stools or urgency?\n"
                "7. Is your hair graying prematurely or thinning?\n"
                "8. Do you feel overheated or intolerant of hot weather?\n"
                "9. Do you get headaches, especially migraines with visual aura?\n"
                "10. Do symptoms worsen in summer or between 10 AM-2 PM?"
            ),
            "aggravating_factors": "Summer/hot weather, midday sun, spicy/sour/fermented food, alcohol, competitive stress, anger, skipping meals",
            "source": "Ashtanga Hridaya Sutrasthana, Charaka Samhita Vol 2, Sushruta Samhita Vol 1",
        },
        {
            "dosha": "Kapha Vikriti",
            "element": "Water + Earth (Jala + Prithvi)",
            "body_type": "Heavy/stocky frame, soft skin, large eyes, thick hair, gains weight easily",
            "key_symptoms": (
                "1. Lethargy, heaviness, excessive sleep\n"
                "2. Weight gain, water retention, edema\n"
                "3. Congestion, mucus, sinus problems\n"
                "4. Depression, attachment, possessiveness\n"
                "5. Slow digestion, loss of appetite\n"
                "6. Oily skin, cystic acne\n"
                "7. Swollen glands, tumors, cysts\n"
                "8. Sweet taste in mouth, excessive salivation"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Light, warm, dry, spicy foods\n"
                "- Pungent, bitter, astringent tastes\n"
                "- Honey (in warm water, not cooked)\n"
                "- Avoid sweet, heavy, oily, cold foods\n\n"
                "HERBS:\n"
                "- Trikatu (ginger, black pepper, pippali)\n"
                "- Guggulu (fat metabolism, detox)\n"
                "- Punarnava (diuretic, reduces swelling)\n"
                "- Bibhitaki (Kapha-specific from Triphala)\n"
                "- Vacha/Calamus (clears congestion)\n\n"
                "THERAPIES:\n"
                "- Vamana (therapeutic emesis - primary Kapha treatment per Charaka)\n"
                "- Udvartana (dry herbal powder massage)\n"
                "- Nasya (nasal administration of herbs)\n\n"
                "ASANAS:\n"
                "- Surya Namaskar (energizing)\n"
                "- Bhastrika Pranayama (bellows breath)\n"
                "- Kapalabhati (skull-shining breath)\n"
                "- Vigorous standing poses\n\n"
                "LIFESTYLE:\n"
                "- Wake before 6 AM (before Kapha time)\n"
                "- Vigorous daily exercise\n"
                "- Avoid daytime sleep\n"
                "- Dry sauna, fasting periodically"
            ),
            "followup_questions": (
                "1. Do you feel heavy, sluggish, or lethargic, especially in the morning?\n"
                "2. Do you gain weight easily and find it hard to lose?\n"
                "3. Do you have sinus congestion, excessive mucus, or frequent colds?\n"
                "4. Do you sleep more than 8 hours and still feel tired?\n"
                "5. Do you experience water retention or swelling?\n"
                "6. Is your appetite low or do you eat out of emotional comfort?\n"
                "7. Do you feel attached, possessive, or resistant to change?\n"
                "8. Do you have oily skin, cystic acne, or benign growths?\n"
                "9. Is your digestion slow — do you feel full long after eating?\n"
                "10. Do symptoms worsen in spring or in cold, damp weather?"
            ),
            "aggravating_factors": "Spring, cold/damp weather, daytime sleep, sweet/heavy/oily food, sedentary lifestyle, overeating, attachment",
            "source": "Ashtanga Hridaya Sutrasthana, Charaka Samhita Vol 2, Sushruta Samhita Vol 1",
        },
    ]

    # ==========================================
    # SHEET 2: Dual Dosha Vikriti
    # ==========================================
    ws2 = wb.create_sheet("Dual Dosha Vikriti")

    dual_dosha_data = [
        {
            "dosha": "Vata-Pitta Vikriti",
            "element": "Air + Fire dominant",
            "body_type": "Thin-medium frame, sensitive skin, sharp mind, nervous energy",
            "key_symptoms": (
                "1. Anxiety WITH irritability\n"
                "2. Burning indigestion with bloating\n"
                "3. Dry skin with occasional rashes\n"
                "4. Insomnia with racing thoughts\n"
                "5. Irregular menstruation with heavy flow\n"
                "6. Joint inflammation (not just cracking)\n"
                "7. Headaches with anxiety\n"
                "8. Alternating constipation and loose stools"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Warm but not hot foods\n"
                "- Sweet taste is best (pacifies both)\n"
                "- Ghee liberally (cools Pitta, nourishes Vata)\n"
                "- Avoid spicy AND raw foods\n\n"
                "HERBS:\n"
                "- Shatavari (cools + nourishes)\n"
                "- Brahmi (calms mind, cools)\n"
                "- Licorice/Yashtimadhu (soothes both)\n"
                "- Ashwagandha (moderate dose)\n\n"
                "THERAPIES:\n"
                "- Mild Abhyanga with Brahmi oil\n"
                "- Shirodhara (calms both doshas)\n"
                "- Mild Basti (address Vata root)\n\n"
                "KEY: Treat Vata first (it drives Pitta out of balance per Ashtanga Hridaya)"
            ),
            "followup_questions": (
                "1. Which came first — the anxiety/restlessness or the anger/burning?\n"
                "2. Does your digestion alternate between gas/bloating and acidity?\n"
                "3. Is your skin both dry AND prone to rashes/inflammation?\n"
                "4. Do you have trouble sleeping AND wake up hot/sweating?\n"
                "5. Are headaches accompanied by anxiety or irritability?\n"
                "6. Does stress manifest as both fear AND frustration?\n"
                "7. Is your energy erratic — wired then crashed?"
            ),
            "aggravating_factors": "Transitional seasons (autumn-summer), stress, irregular schedule + hot food, fasting",
            "source": "Ashtanga Hridaya Sutrasthana Ch.13, Sushruta Samhita Vol 1",
        },
        {
            "dosha": "Vata-Kapha Vikriti",
            "element": "Air + Earth dominant (Cold is common factor)",
            "body_type": "Variable frame, cold body, congestion-prone with anxiety",
            "key_symptoms": (
                "1. Anxiety WITH depression/lethargy\n"
                "2. Cold body with congestion\n"
                "3. Constipation with mucus in stool\n"
                "4. Weight gain despite poor appetite\n"
                "5. Joint stiffness with cracking\n"
                "6. Brain fog with restless thoughts\n"
                "7. Respiratory issues (asthma, wheezing)\n"
                "8. Bloating with slow digestion"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Warm, light, spiced foods\n"
                "- Pungent taste balances both\n"
                "- Ginger tea throughout the day\n"
                "- Avoid cold, heavy, raw, sweet foods\n\n"
                "HERBS:\n"
                "- Trikatu (warms, clears, stimulates)\n"
                "- Chitrak (digestive fire)\n"
                "- Sitopaladi (respiratory)\n"
                "- Pippali (long pepper - best for Vata-Kapha)\n\n"
                "THERAPIES:\n"
                "- Udvartana (stimulating dry massage)\n"
                "- Swedana (herbal steam therapy)\n"
                "- Nasya (clears sinuses)\n\n"
                "KEY: Build Agni (digestive fire) — weak Agni is the root of Vata-Kapha imbalance"
            ),
            "followup_questions": (
                "1. Do you feel both anxious AND heavy/sluggish at the same time?\n"
                "2. Are you always cold, regardless of the weather?\n"
                "3. Do you have both constipation AND mucus/congestion?\n"
                "4. Is your appetite very low but weight still increases?\n"
                "5. Do you have breathing difficulties, asthma, or wheezing?\n"
                "6. Are your joints both stiff AND cracking?\n"
                "7. Do you feel mentally foggy yet unable to calm your thoughts?"
            ),
            "aggravating_factors": "Cold/damp seasons (late winter-spring), cold food, sedentary life + irregular schedule",
            "source": "Ashtanga Hridaya Sutrasthana, Charaka Samhita Vol 2",
        },
        {
            "dosha": "Pitta-Kapha Vikriti",
            "element": "Fire + Water dominant (Liquid/oily is common factor)",
            "body_type": "Medium-heavy frame, oily skin, prone to inflammation + congestion",
            "key_symptoms": (
                "1. Anger WITH stubbornness\n"
                "2. Oily skin with inflamed acne\n"
                "3. Excess mucus that is yellow/green (infected)\n"
                "4. Weight gain with inflammation\n"
                "5. Diabetes tendency (Prameha)\n"
                "6. Fatty liver, high cholesterol\n"
                "7. Sinus infections (not just congestion)\n"
                "8. Excessive sweating with congestion"
            ),
            "immediate_cure": (
                "DIET:\n"
                "- Bitter + astringent tastes (balance both)\n"
                "- Light, dry, cooling but not cold\n"
                "- Green vegetables, turmeric, bitter gourd\n"
                "- Avoid sweet, oily, fried, sour foods\n\n"
                "HERBS:\n"
                "- Kutki (liver, cools, reduces Kapha)\n"
                "- Neem (blood purifier, bitter)\n"
                "- Turmeric (anti-inflammatory, drying)\n"
                "- Guggulu (fat metabolism)\n"
                "- Triphala (balances all three)\n\n"
                "THERAPIES:\n"
                "- Virechana (purgation for Pitta)\n"
                "- Udvartana (dry massage for Kapha)\n"
                "- Raktamokshana (bloodletting per Sushruta for inflammatory conditions)\n\n"
                "KEY: Address Ama (toxins) first — Pitta-Kapha creates toxic accumulation"
            ),
            "followup_questions": (
                "1. Is your mucus/congestion yellow or green (indicating Pitta involvement)?\n"
                "2. Do you have oily skin with inflamed (not just cystic) acne?\n"
                "3. Are you prone to infections, especially sinus or urinary?\n"
                "4. Do you have high cholesterol, fatty liver, or metabolic issues?\n"
                "5. Do you experience both overheating AND water retention?\n"
                "6. Is your anger accompanied by rigid/stubborn thinking?\n"
                "7. Do you have any blood sugar concerns or Prameha symptoms?"
            ),
            "aggravating_factors": "Late spring-summer, humid weather, oily/sweet/sour food, sedentary + stressful lifestyle",
            "source": "Ashtanga Hridaya Sutrasthana, Sushruta Samhita Vol 2",
        },
    ]

    # ==========================================
    # SHEET 3: Tridosha Vikriti (Sannipata)
    # ==========================================
    ws3 = wb.create_sheet("Tridosha Vikriti (Sannipata)")

    tridosha_data = [
        {
            "dosha": "Sannipata (Tridosha Vikriti)",
            "element": "All five elements disturbed",
            "body_type": "Variable — symptoms are contradictory and shifting",
            "key_symptoms": (
                "1. Contradictory symptoms (fever + chills, diarrhea + constipation)\n"
                "2. Rapidly changing condition\n"
                "3. Severe weakness and depletion\n"
                "4. Mental confusion, delirium\n"
                "5. Multiple organ systems involved\n"
                "6. Poor prognosis if untreated (per Sushruta)\n"
                "7. Unstable vital signs\n"
                "8. Loss of taste, smell, or sensation"
            ),
            "immediate_cure": (
                "CRITICAL: Sannipata requires expert Vaidya supervision.\n\n"
                "APPROACH:\n"
                "- Identify the PREDOMINANT dosha and treat it first\n"
                "- Avoid strong Panchakarma initially (patient too weak)\n"
                "- Build strength (Brimhana) before purification\n\n"
                "HERBS:\n"
                "- Triphala (balances all three doshas)\n"
                "- Dashamoola (10-root formula, broad spectrum)\n"
                "- Chyawanprash (rejuvenative, safe for all)\n"
                "- Guduchi (immune support, tridoshic)\n\n"
                "DIET:\n"
                "- Light, warm, easily digestible\n"
                "- Yusha (medicated soups per Charaka)\n"
                "- Rice gruel (Peya/Manda)\n"
                "- Small, frequent meals\n\n"
                "THERAPIES:\n"
                "- Gentle external oil application only initially\n"
                "- Progress to mild Basti once strength returns\n"
                "- Sequential dosha treatment (predominant first)"
            ),
            "followup_questions": (
                "1. Are your symptoms contradictory (e.g., feeling hot and cold simultaneously)?\n"
                "2. How rapidly do your symptoms change — hours, days?\n"
                "3. How long has this multi-system imbalance been present?\n"
                "4. What is your current strength level (can you walk, eat, sleep)?\n"
                "5. Which symptom is MOST dominant right now?\n"
                "6. Have you had any recent severe illness, surgery, or trauma?\n"
                "7. Are you experiencing mental confusion or disorientation?\n"
                "8. What is your current diet and medication?\n"
                "9. Is there a seasonal pattern to the worsening?\n"
                "10. URGENT: Are you under the care of a qualified Ayurvedic practitioner?"
            ),
            "aggravating_factors": "Seasonal junctions (Ritu Sandhi), prolonged illness, improper treatment of single/dual dosha, severe stress + poor diet + no routine",
            "source": "Sushruta Samhita Vol 2 (Sannipata), Charaka Samhita, Ashtanga Hridaya Sutrasthana",
        },
    ]

    # ==========================================
    # STYLING
    # ==========================================
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    vata_fill = PatternFill(start_color="E8D5B7", end_color="E8D5B7", fill_type="solid")
    pitta_fill = PatternFill(start_color="F4C2C2", end_color="F4C2C2", fill_type="solid")
    kapha_fill = PatternFill(start_color="B7D7E8", end_color="B7D7E8", fill_type="solid")
    dual_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    tri_fill = PatternFill(start_color="E8D4E8", end_color="E8D4E8", fill_type="solid")

    dosha_fills = {
        "Vata Vikriti": vata_fill,
        "Pitta Vikriti": pitta_fill,
        "Kapha Vikriti": kapha_fill,
        "Vata-Pitta Vikriti": dual_fill,
        "Vata-Kapha Vikriti": dual_fill,
        "Pitta-Kapha Vikriti": dual_fill,
        "Sannipata (Tridosha Vikriti)": tri_fill,
    }

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Vikriti Type",
        "Elements",
        "Body Type / Characteristics",
        "Key Symptoms",
        "Immediate Cure (Diet, Herbs, Therapies, Asanas)",
        "Follow-up Diagnostic Questions",
        "Aggravating Factors",
        "Classical Source Reference",
    ]

    col_widths = [20, 18, 30, 40, 55, 50, 35, 30]

    def write_sheet(ws, data, title):
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Headers
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 30

        # Data rows
        for row_idx, d in enumerate(data, start=3):
            row_data = [
                d["dosha"],
                d["element"],
                d["body_type"],
                d["key_symptoms"],
                d["immediate_cure"],
                d["followup_questions"],
                d["aggravating_factors"],
                d["source"],
            ]

            fill = dosha_fills.get(d["dosha"], None)

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = wrap_alignment
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                if col == 1:
                    cell.font = Font(bold=True, size=11)

            ws.row_dimensions[row_idx].height = 350

    write_sheet(ws1, single_dosha_data, "SINGLE DOSHA VIKRITI — Symptoms, Cures & Diagnostic Questions")
    write_sheet(ws2, dual_dosha_data, "DUAL DOSHA VIKRITI — Symptoms, Cures & Diagnostic Questions")
    write_sheet(ws3, tridosha_data, "TRIDOSHA VIKRITI (SANNIPATA) — Symptoms, Cures & Diagnostic Questions")

    # ==========================================
    # SHEET 4: Quick Diagnostic Flowchart
    # ==========================================
    ws4 = wb.create_sheet("Diagnostic Flowchart")
    ws4.merge_cells("A1:E1")
    title = ws4.cell(row=1, column=1, value="VIKRITI DIAGNOSTIC FLOWCHART — Key Differentiating Questions")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
    title.alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 35

    flow_headers = ["Step", "Question to Ask", "If YES → Indicates", "If NO → Proceed to", "Notes"]
    for col, h in enumerate(flow_headers, 1):
        cell = ws4.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.border = thin_border

    flow_data = [
        ["1", "Is the patient feeling cold, dry, and anxious?", "Vata involvement", "Step 2", "Check extremities temperature, skin texture"],
        ["2", "Is there heat, inflammation, or burning sensation?", "Pitta involvement", "Step 3", "Check eyes, skin, stomach acid"],
        ["3", "Is there heaviness, congestion, or excess mucus?", "Kapha involvement", "Step 4", "Check weight, sinuses, energy level"],
        ["4", "Are symptoms contradictory or rapidly changing?", "Sannipata (Tridosha)", "Step 5", "REFER to experienced Vaidya immediately"],
        ["5", "Do cold AND heat symptoms coexist?", "Vata-Pitta dual", "Step 6", "Anxiety + acidity = classic Vata-Pitta"],
        ["6", "Do cold AND heaviness coexist?", "Vata-Kapha dual", "Step 7", "Anxiety + congestion = Vata-Kapha"],
        ["7", "Do heat AND heaviness coexist?", "Pitta-Kapha dual", "Single dosha", "Inflammation + mucus = Pitta-Kapha"],
        ["8", "When did symptoms start / what season?", "Seasonal correlation", "Step 9", "Vata=autumn, Pitta=summer, Kapha=spring"],
        ["9", "What time of day are symptoms worst?", "Time correlation", "Step 10", "Vata=2-6, Pitta=10-2, Kapha=6-10"],
        ["10", "What is the patient's Prakriti (birth constitution)?", "Baseline comparison", "Treatment plan", "Vikriti = deviation from Prakriti"],
    ]

    for row_idx, row in enumerate(flow_data, start=3):
        for col, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col, value=value)
            cell.alignment = wrap_alignment
            cell.border = thin_border

    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 45
    ws4.column_dimensions["C"].width = 25
    ws4.column_dimensions["D"].width = 22
    ws4.column_dimensions["E"].width = 45

    for r in range(3, 13):
        ws4.row_dimensions[r].height = 35

    # Save
    output = "data/Vikriti_Dosha_Diagnosis_Guide.xlsx"
    wb.save(output)
    print(f"Saved to {output}")
    return output


if __name__ == "__main__":
    create_vikriti_sheet()
