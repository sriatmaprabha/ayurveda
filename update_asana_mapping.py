"""Add Asana-Condition mapping sheet + Asana details sheet to the diagnosis Excel."""

import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load data
with open("data/asana_full_details.json", "r", encoding="utf-8") as f:
    sql_asanas = json.load(f)

with open("data/asana_protocol_analysis.json", "r", encoding="utf-8") as f:
    protocol_data = json.load(f)

with open("data/asana_protocols.json", "r", encoding="utf-8") as f:
    raw_protocols = json.load(f)

wb = load_workbook("data/Prakriti_Dosha_Diagnosis_Tree.xlsx")

# === STYLING ===
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
asana_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
kumbhaka_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
protocol_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# =====================================================================
# SHEET 5: TOP ASANAS WITH FULL INSTRUCTIONS
# =====================================================================
ws5 = wb.create_sheet("Asana Details (Top 50)")

ws5.merge_cells("A1:H1")
title = ws5.cell(row=1, column=1, value="TOP 50 MOST PRESCRIBED ASANAS & KUMBHAKAS -- With Full Instructions from insert_asanas.sql")
title.font = Font(bold=True, size=13, color="FFFFFF")
title.fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
title.alignment = Alignment(horizontal="center")
ws5.row_dimensions[1].height = 30

headers5 = ["#", "Asana/Kriya Name", "Sanskrit Name", "Type", "Difficulty",
            "Full Technique / How to Perform", "Benefits", "Used in Protocols (Count)"]
widths5 = [5, 25, 20, 12, 12, 55, 45, 35]

for col, (h, w) in enumerate(zip(headers5, widths5), 1):
    cell = ws5.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws5.column_dimensions[get_column_letter(col)].width = w

# Get frequency data
freq = protocol_data["asana_frequency"]

# Build list: most-used asanas that have details in SQL
asana_rows = []
for asana_name, freq_data in sorted(freq.items(), key=lambda x: x[1]["count"], reverse=True):
    # Try to find in SQL details
    detail = None
    for sql_name, sql_detail in sql_asanas.items():
        if (asana_name.lower() in sql_name.lower() or
            sql_name.lower() in asana_name.lower() or
            sql_detail.get("sanskrit_name", "").lower() in asana_name.lower()):
            detail = sql_detail
            break

    asana_type = "Kumbhaka/Pranayama" if "KUMBHAKA" in asana_name or "PRANAYAMA" in asana_name or "PRAANAAYAAMA" in asana_name else "Asana/Bandha"

    protocols_list = freq_data["protocols"][:8]

    if detail:
        asana_rows.append({
            "name": asana_name.title(),
            "sanskrit": detail.get("sanskrit_name", asana_name),
            "type": asana_type,
            "difficulty": detail.get("difficulty", "intermediate"),
            "technique": detail.get("technique", "See protocol details"),
            "benefits": detail.get("benefits", ""),
            "count": freq_data["count"],
            "protocols": ", ".join(protocols_list),
        })
    else:
        asana_rows.append({
            "name": asana_name.title(),
            "sanskrit": asana_name.title(),
            "type": asana_type,
            "difficulty": "intermediate",
            "technique": "(Technique in protocol step details -- see Condition-Asana Map sheet)",
            "benefits": "",
            "count": freq_data["count"],
            "protocols": ", ".join(protocols_list),
        })

    if len(asana_rows) >= 50:
        break

row = 3
for i, a in enumerate(asana_rows, 1):
    data = [i, a["name"], a["sanskrit"], a["type"], a["difficulty"],
            a["technique"], a["benefits"], f"{a['count']} protocols:\n{a['protocols']}"]

    fill = kumbhaka_fill if "Kumbhaka" in a["type"] or "Pranayama" in a["type"] else asana_fill

    for col, val in enumerate(data, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        cell.alignment = wrap
        cell.border = border
        if col in (2, 6):
            cell.fill = fill
        if col == 2:
            cell.font = Font(bold=True)

    ws5.row_dimensions[row].height = 120
    row += 1

# =====================================================================
# SHEET 6: CONDITION → ASANA FULL MAP
# =====================================================================
ws6 = wb.create_sheet("Condition-Asana Map")

ws6.merge_cells("A1:F1")
title = ws6.cell(row=1, column=1, value="CONDITION → ASANA PROTOCOL MAP -- 112 Protocols with Full Step-by-Step Instructions")
title.font = Font(bold=True, size=13, color="FFFFFF")
title.fill = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
title.alignment = Alignment(horizontal="center")
ws6.row_dimensions[1].height = 30

headers6 = ["Condition (Care/Cure)", "# of Asanas", "Asana Sequence",
            "Full Step-by-Step Instructions", "Type (Care/Cure)", "Dosha Relevance"]
widths6 = [30, 10, 35, 70, 12, 15]

for col, (h, w) in enumerate(zip(headers6, widths6), 1):
    cell = ws6.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws6.column_dimensions[get_column_letter(col)].width = w

# Dosha mapping for protocols
condition_dosha = {
    "Anxiety": "Vata", "Anger": "Pitta", "Arthritis": "Vata",
    "Asthma": "Kapha", "Depression": "Kapha", "Diabetes": "Kapha",
    "Obesity": "Kapha", "Insomnia": "Vata", "Hypertension": "Pitta",
    "Migraine": "Pitta", "Sinusitis": "Kapha", "Skin": "Pitta",
    "Heart": "Pitta", "Back Pain": "Vata", "Lower Back Pain": "Vata",
    "Thyroid": "Kapha", "Memory": "Vata", "Epilepsy": "Vata",
    "Vertigo": "Vata", "Tinnitus": "Vata", "Baldness": "Pitta",
    "Addiction": "Vata", "Cancer": "All", "Eczema": "Pitta",
    "Infection": "Pitta", "Hernia": "Vata", "Ageing": "Vata",
    "Autism": "Vata", "Bipolar": "Vata-Pitta",
    "Schizophrenia": "Vata", "Dandruff": "Pitta-Kapha",
    "Fungal": "Kapha", "Food Allerg": "Pitta",
    "Stiff Knees": "Vata", "Kidney": "Kapha",
    "Nephrotic": "Kapha", "Pulmonary": "Kapha",
    "Short-Sight": "Pitta", "Long Sight": "Pitta",
    "Colitis": "Pitta", "Urinary": "Vata-Kapha",
    "Self Esteem": "Kapha", "Sweating": "Pitta",
    "Hot Flashes": "Pitta", "Polycystic": "Kapha",
    "Achromatopsia": "Pitta", "Hypothyroid": "Kapha",
}

row = 3
for pname in sorted(raw_protocols.keys()):
    if pname == "Name of kriya":
        continue

    pdata = raw_protocols[pname]
    steps_summary = pdata.get("steps_summary", "")
    step_details = pdata.get("step_details", [])

    if not steps_summary and not step_details:
        continue

    # Count asanas
    asana_count = len([l for l in steps_summary.split("\n") if l.strip() and l.strip()[0].isdigit()])
    if asana_count == 0:
        asana_count = len(step_details)

    # Full instructions
    full_instructions = ""
    for i, step in enumerate(step_details, 1):
        full_instructions += f"STEP {i}:\n{step.strip()}\n\n"

    # Protocol type
    ptype = "Care" if pname.lower().startswith("care") else "Cure" if pname.lower().startswith("cure") else "Other"

    # Dosha
    dosha = "General"
    for keyword, d in condition_dosha.items():
        if keyword.lower() in pname.lower():
            dosha = d
            break

    fill = asana_fill if ptype == "Care" else kumbhaka_fill if ptype == "Cure" else protocol_fill

    data = [pname, asana_count, steps_summary, full_instructions, ptype, dosha]

    for col, val in enumerate(data, 1):
        cell = ws6.cell(row=row, column=col, value=val)
        cell.alignment = wrap
        cell.border = border
        if col == 1:
            cell.font = Font(bold=True)
        if col == 5:
            cell.fill = fill

    ws6.row_dimensions[row].height = 350 if full_instructions else 80
    row += 1

# =====================================================================
# ADD SUMMARY STATS ROW
# =====================================================================
row += 1
ws6.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
row += 1
stats = [
    f"Total Protocols: {len(raw_protocols) - 1}",
    f"Total Unique Asanas/Kumbhakas: {protocol_data['total_unique_asanas']}",
    f"Asana Details (from SQL): {len(sql_asanas)} poses with full technique",
    f"Scripture References: 15 books mention yoga/asana",
    f"Most Used: Trinetra Kumbhaka (20 protocols), Trishoola Kumbhaka (18), Naaga Kumbhaka (17)",
]
for s in stats:
    ws6.cell(row=row, column=1, value=s)
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

try:
    wb.save("data/Prakriti_Dosha_Diagnosis_Tree.xlsx")
except PermissionError:
    wb.save("data/Prakriti_Dosha_Diagnosis_Tree_v2.xlsx")
    print("  (Saved as v2 -- original file was open in Excel)")
print("Updated Excel with 2 new sheets:")
print(f"  Sheet 5: Top 50 Asanas with full instructions ({len(asana_rows)} entries)")
print(f"  Sheet 6: Condition-Asana Map ({len(raw_protocols)-1} protocols with full steps)")
